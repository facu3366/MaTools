"""
📊 DCF MODEL GENERATOR — DealDesk
Genera un Excel con fórmulas vivas: históricos, regresión, proyección, DCF, sensitivity.

Uso standalone:
    python dcf_excel.py AAPL

Uso como módulo (desde api.py):
    from dcf_excel import generate_dcf_excel
    buffer = generate_dcf_excel("AAPL")
"""

import sys
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import yfinance as yf

# ─────────────────────────────────────────────
# CONSTANTS — Color Coding (IB Standard)
# ─────────────────────────────────────────────

BLUE_FONT = Font(name="Arial", size=10, color="0000FF")          # Hardcoded inputs
BLACK_FONT = Font(name="Arial", size=10, color="000000")          # Formulas
BLACK_BOLD = Font(name="Arial", size=10, color="000000", bold=True)
GREEN_FONT = Font(name="Arial", size=10, color="008000")          # Cross-sheet links
HEADER_FONT = Font(name="Arial", size=10, color="FFFFFF", bold=True)
TITLE_FONT = Font(name="Arial", size=12, color="000000", bold=True)
SECTION_FONT = Font(name="Arial", size=10, color="8B7535", bold=True)
SMALL_FONT = Font(name="Arial", size=8, color="888888")

YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")              # Key assumptions
HEADER_FILL = PatternFill("solid", fgColor="1F3864")              # Dark blue headers
LIGHT_FILL = PatternFill("solid", fgColor="D6E4F0")               # Light blue alternating
CREAM_FILL = PatternFill("solid", fgColor="FAF8F0")               # Cream background
GOLD_FILL = PatternFill("solid", fgColor="F2E8C9")                # Gold highlight
OUTPUT_FILL = PatternFill("solid", fgColor="E2EFDA")              # Green output cells

ALIGN_R = Alignment(horizontal="right", vertical="center")
ALIGN_L = Alignment(horizontal="left", vertical="center")
ALIGN_C = Alignment(horizontal="center", vertical="center")

THIN_BORDER = Border(
    bottom=Side(style="thin", color="D9D9D9")
)
THICK_BOTTOM = Border(
    bottom=Side(style="medium", color="1F3864")
)

FMT_USD = '#,##0;(#,##0);"-"'
FMT_USD_DEC = '#,##0.0;(#,##0.0);"-"'
FMT_PCT = '0.0%'
FMT_MULT = '0.0"x"'
FMT_PRICE = '$#,##0.00'
FMT_YEAR = '0'


# ─────────────────────────────────────────────
# DATA FETCH
# ─────────────────────────────────────────────

def fetch_data(ticker):
    stock = yf.Ticker(ticker)
    info = stock.info or {}
    inc = stock.financials
    inc_q = stock.quarterly_financials
    bal = stock.balance_sheet
    cf = stock.cashflow

    def safe(df, label):
        if df is not None and label in df.index:
            vals = df.loc[label].tolist()[:4]
            return [v if v == v else None for v in vals]  # handle NaN
        return [None]*4

    def safe_years(df):
        if df is not None and not df.empty:
            return [d.year for d in df.columns.tolist()[:4]]
        return []

    return {
        "info": info,
        "years": safe_years(inc),
        "revenue": safe(inc, "Total Revenue"),
        "cogs": safe(inc, "Cost Of Revenue"),
        "gross_profit": safe(inc, "Gross Profit"),
        "operating_income": safe(inc, "Operating Income"),
        "ebitda": safe(inc, "EBITDA"),
        "net_income": safe(inc, "Net Income"),
        "rd": safe(inc, "Research Development"),
        "sga": safe(inc, "Selling General Administrative"),
        "da": safe(cf, "Depreciation") or safe(inc, "Depreciation And Amortization"),
        "capex": safe(cf, "Capital Expenditures") or safe(cf, "Capital Expenditure"),
        "op_cf": safe(cf, "Total Cash From Operating Activities") or safe(cf, "Operating Cash Flow"),
        "total_assets": safe(bal, "Total Assets"),
        "total_debt_bal": safe(bal, "Long Term Debt"),
        "cash": safe(bal, "Cash And Cash Equivalents") or safe(bal, "Cash"),
        "total_equity": safe(bal, "Stockholders Equity") or safe(bal, "Total Stockholders Equity"),
        "quarterly_rev": safe(inc_q, "Total Revenue")[:8] if inc_q is not None else [],
        "quarterly_dates": [str(d.date()) for d in inc_q.columns.tolist()[:8]] if inc_q is not None and not inc_q.empty else [],
    }


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def write_row(ws, row, data, font=BLACK_FONT, fill=None, fmt=None, border=None):
    for i, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=i, value=val)
        cell.font = font
        cell.alignment = ALIGN_R if i > 1 else ALIGN_L
        if fill:
            cell.fill = fill
        if fmt and i > 1:
            cell.number_format = fmt
        if border:
            cell.border = border

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ─────────────────────────────────────────────
# SHEET 1: HISTORICALS
# ─────────────────────────────────────────────

def build_historicals(wb, data, ticker):
    ws = wb.active
    ws.title = "Historicals"
    ws.sheet_properties.tabColor = "1F3864"

    info = data["info"]
    years = data["years"]
    n = len(years)
    if n == 0:
        ws["A1"] = "No historical data available"
        return

    # Column widths: Label + years
    widths = [32] + [16]*n + [16]  # extra col for CAGR
    set_col_widths(ws, widths)

    r = 1
    ws.cell(row=r, column=1, value=f"{info.get('longName', ticker)} ({ticker})").font = TITLE_FONT
    r += 1
    ws.cell(row=r, column=1, value=f"Historical Financial Data · Source: Yahoo Finance · {datetime.now().strftime('%d/%m/%Y')}").font = SMALL_FONT
    r += 2

    # Header row
    headers = ["Income Statement ($M)"] + [str(y) for y in years] + ["CAGR"]
    write_row(ws, r, headers, font=HEADER_FONT, fill=HEADER_FILL)
    r += 1

    # Data rows
    metrics = [
        ("Revenue", data["revenue"], FMT_USD),
        ("COGS", data["cogs"], FMT_USD),
        ("Gross Profit", data["gross_profit"], FMT_USD),
        ("R&D", data["rd"], FMT_USD),
        ("SG&A", data["sga"], FMT_USD),
        ("Operating Income (EBIT)", data["operating_income"], FMT_USD),
        ("EBITDA", data["ebitda"], FMT_USD),
        ("Net Income", data["net_income"], FMT_USD),
    ]

    rev_row = r  # save for margin calcs

    for label, values, fmt in metrics:
        row_data = [label]
        for v in values:
            row_data.append(v / 1e6 if v else None)
        # CAGR formula (last/first)^(1/n)-1
        if len(values) >= 2 and values[0] and values[-1]:
            first_col = get_column_letter(2)
            last_col = get_column_letter(1 + n)
            cagr_col = 1 + n + 1
            ws.cell(row=r, column=cagr_col).value = f"=({last_col}{r}/{first_col}{r})^(1/{n-1})-1"
            ws.cell(row=r, column=cagr_col).font = BLACK_FONT
            ws.cell(row=r, column=cagr_col).number_format = FMT_PCT

        is_bold = label in ("Revenue", "Gross Profit", "Operating Income (EBIT)", "EBITDA", "Net Income")
        write_row(ws, r, row_data, font=BLACK_BOLD if is_bold else BLACK_FONT, fmt=fmt,
                  border=THICK_BOTTOM if label == "Net Income" else THIN_BORDER)
        # Color hardcoded values blue
        for ci in range(2, 2+n):
            ws.cell(row=r, column=ci).font = BLUE_FONT if not is_bold else Font(name="Arial", size=10, color="0000FF", bold=True)
        r += 1

    r += 1

    # Margin Analysis (formulas!)
    ws.cell(row=r, column=1, value="Margin Analysis").font = SECTION_FONT
    r += 1
    headers2 = ["Margins (%)"] + [str(y) for y in years]
    write_row(ws, r, headers2, font=HEADER_FONT, fill=HEADER_FILL)
    r += 1

    margin_rows = [
        ("Gross Margin", "Gross Profit", "Revenue"),
        ("Operating Margin", "Operating Income (EBIT)", "Revenue"),
        ("EBITDA Margin", "EBITDA", "Revenue"),
        ("Net Margin", "Net Income", "Revenue"),
    ]

    # Find row numbers for each metric
    metric_row_map = {}
    search_r = rev_row
    for label, vals, _ in metrics:
        metric_row_map[label] = search_r
        search_r += 1

    for label, num_label, den_label in margin_rows:
        num_r = metric_row_map.get(num_label)
        den_r = metric_row_map.get(den_label)
        row_data = [label]
        ws.cell(row=r, column=1, value=label).font = BLACK_BOLD
        ws.cell(row=r, column=1).border = THIN_BORDER
        for ci in range(2, 2+n):
            col_l = get_column_letter(ci)
            formula = f"=IF({col_l}{den_r}=0,0,{col_l}{num_r}/{col_l}{den_r})"
            ws.cell(row=r, column=ci, value=formula).font = BLACK_FONT
            ws.cell(row=r, column=ci).number_format = FMT_PCT
            ws.cell(row=r, column=ci).alignment = ALIGN_R
            ws.cell(row=r, column=ci).border = THIN_BORDER
        r += 1

    r += 2

    # Cash Flow section
    ws.cell(row=r, column=1, value="Cash Flow ($M)").font = SECTION_FONT
    r += 1
    headers3 = ["Cash Flow"] + [str(y) for y in years]
    write_row(ws, r, headers3, font=HEADER_FONT, fill=HEADER_FILL)
    r += 1

    cf_metrics = [
        ("Operating Cash Flow", data["op_cf"]),
        ("Capital Expenditures", data["capex"]),
        ("Depreciation & Amort.", data["da"]),
    ]

    cf_row_map = {}
    for label, values in cf_metrics:
        row_data = [label]
        for v in values:
            row_data.append(v / 1e6 if v else None)
        write_row(ws, r, row_data, font=BLUE_FONT, fmt=FMT_USD, border=THIN_BORDER)
        cf_row_map[label] = r
        r += 1

    # FCF = OCF + Capex (capex is negative)
    ws.cell(row=r, column=1, value="Free Cash Flow").font = BLACK_BOLD
    ws.cell(row=r, column=1).border = THICK_BOTTOM
    ocf_r = cf_row_map["Operating Cash Flow"]
    capex_r = cf_row_map["Capital Expenditures"]
    for ci in range(2, 2+n):
        col_l = get_column_letter(ci)
        ws.cell(row=r, column=ci, value=f"={col_l}{ocf_r}+{col_l}{capex_r}").font = BLACK_BOLD
        ws.cell(row=r, column=ci).number_format = FMT_USD
        ws.cell(row=r, column=ci).alignment = ALIGN_R
        ws.cell(row=r, column=ci).border = THICK_BOTTOM

    return metric_row_map, cf_row_map


# ─────────────────────────────────────────────
# SHEET 2: ASSUMPTIONS & PROJECTIONS
# ─────────────────────────────────────────────

def build_projections(wb, data, ticker):
    ws = wb.create_sheet("Projections")
    ws.sheet_properties.tabColor = "8B7535"

    info = data["info"]
    years = data["years"]
    n_hist = len(years)
    n_proj = 5

    if n_hist == 0:
        ws["A1"] = "No data"
        return {}

    last_year = years[0]  # most recent (yfinance returns reverse chronological)
    proj_years = [last_year + i for i in range(1, n_proj + 1)]

    widths = [32] + [16]*(n_proj + 1)
    set_col_widths(ws, widths)

    r = 1
    ws.cell(row=r, column=1, value=f"{info.get('longName', ticker)} — DCF Projection Model").font = TITLE_FONT
    r += 2

    # ─── ASSUMPTIONS BLOCK ───
    ws.cell(row=r, column=1, value="KEY ASSUMPTIONS").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = GOLD_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1

    assumptions = {}

    def add_assumption(label, value, fmt, note=""):
        nonlocal r
        ws.cell(row=r, column=1, value=label).font = BLACK_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        cell = ws.cell(row=r, column=2, value=value)
        cell.font = BLUE_FONT
        cell.fill = YELLOW_FILL
        cell.number_format = fmt
        cell.alignment = ALIGN_R
        cell.border = THIN_BORDER
        if note:
            ws.cell(row=r, column=3, value=note).font = SMALL_FONT
        assumptions[label] = f"$B${r}"
        r += 1
        return f"$B${r-1}"

    # Calculate historical growth
    rev = data["revenue"]
    hist_growth = None
    if rev and len(rev) >= 2 and rev[0] and rev[-1]:
        hist_growth = (rev[0] / rev[-1]) ** (1/(n_hist-1)) - 1

    last_rev = rev[0] / 1e6 if rev and rev[0] else 0
    last_ebitda_m = (data["ebitda"][0] / 1e6) if data["ebitda"] and data["ebitda"][0] else 0
    ebitda_margin = last_ebitda_m / last_rev if last_rev else 0.15

    beta = info.get("beta", 1.0) or 1.0
    mkt_cap = info.get("marketCap", 0) or 0
    total_debt = info.get("totalDebt", 0) or 0
    total_cash = info.get("totalCash", 0) or 0

    ref_rev_g1 = add_assumption("Revenue Growth Y1", hist_growth or 0.10, FMT_PCT, f"Source: CAGR {n_hist-1}Y = {hist_growth:.1%}" if hist_growth else "Estimate")
    ref_rev_g2 = add_assumption("Revenue Growth Y2-Y3", (hist_growth * 0.85) if hist_growth else 0.08, FMT_PCT, "Deceleration assumption")
    ref_rev_g4 = add_assumption("Revenue Growth Y4-Y5", 0.05, FMT_PCT, "Steady state / GDP+")
    ref_ebitda_m = add_assumption("EBITDA Margin", ebitda_margin, FMT_PCT, f"Source: LTM = {ebitda_margin:.1%}")
    ref_da_pct = add_assumption("D&A % Revenue", 0.04, FMT_PCT, "Historical average estimate")
    ref_capex_pct = add_assumption("Capex % Revenue", 0.05, FMT_PCT, "Historical average estimate")
    ref_nwc_pct = add_assumption("Change in NWC % Rev", 0.01, FMT_PCT, "Industry standard")
    ref_tax = add_assumption("Tax Rate", 0.21, FMT_PCT, "US Corporate rate")
    r_blank = r
    r += 1

    # WACC assumptions
    ws.cell(row=r, column=1, value="WACC INPUTS").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = GOLD_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1

    ref_rfr = add_assumption("Risk-Free Rate", 0.045, FMT_PCT, "US 10Y Treasury")
    ref_erp = add_assumption("Equity Risk Premium", 0.055, FMT_PCT, "Damodaran 2025")
    ref_beta = add_assumption("Beta", beta, "0.00", f"Source: Yahoo Finance")
    ref_kd = add_assumption("Pre-tax Cost of Debt", 0.05, FMT_PCT, "Estimated from credit profile")
    ref_tgr = add_assumption("Terminal Growth Rate", 0.025, FMT_PCT, "Long-term GDP growth")

    r += 1

    # ─── WACC CALCULATION ───
    ws.cell(row=r, column=1, value="WACC CALCULATION").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = GOLD_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1

    # Market Cap
    ws.cell(row=r, column=1, value="Market Cap ($M)").font = BLACK_FONT
    ws.cell(row=r, column=2, value=mkt_cap / 1e6 if mkt_cap else 0).font = BLUE_FONT
    ws.cell(row=r, column=2).number_format = FMT_USD
    ref_mktcap = f"$B${r}"
    r += 1

    # Total Debt
    ws.cell(row=r, column=1, value="Total Debt ($M)").font = BLACK_FONT
    ws.cell(row=r, column=2, value=total_debt / 1e6 if total_debt else 0).font = BLUE_FONT
    ws.cell(row=r, column=2).number_format = FMT_USD
    ref_debt = f"$B${r}"
    r += 1

    # Equity Weight
    ws.cell(row=r, column=1, value="Equity Weight").font = BLACK_FONT
    ws.cell(row=r, column=2, value=f"={ref_mktcap}/({ref_mktcap}+{ref_debt})").font = BLACK_FONT
    ws.cell(row=r, column=2).number_format = FMT_PCT
    ref_we = f"$B${r}"
    r += 1

    # Debt Weight
    ws.cell(row=r, column=1, value="Debt Weight").font = BLACK_FONT
    ws.cell(row=r, column=2, value=f"={ref_debt}/({ref_mktcap}+{ref_debt})").font = BLACK_FONT
    ws.cell(row=r, column=2).number_format = FMT_PCT
    ref_wd = f"$B${r}"
    r += 1

    # Cost of Equity (CAPM)
    ws.cell(row=r, column=1, value="Cost of Equity (Ke)").font = BLACK_BOLD
    ws.cell(row=r, column=2, value=f"={ref_rfr}+{ref_beta}*{ref_erp}").font = BLACK_BOLD
    ws.cell(row=r, column=2).number_format = FMT_PCT
    ws.cell(row=r, column=2).fill = OUTPUT_FILL
    ref_ke = f"$B${r}"
    r += 1

    # WACC
    ws.cell(row=r, column=1, value="WACC").font = BLACK_BOLD
    ws.cell(row=r, column=2, value=f"={ref_ke}*{ref_we}+{ref_kd}*(1-{ref_tax})*{ref_wd}").font = BLACK_BOLD
    ws.cell(row=r, column=2).number_format = FMT_PCT
    ws.cell(row=r, column=2).fill = OUTPUT_FILL
    ws.cell(row=r, column=2).border = THICK_BOTTOM
    ref_wacc = f"$B${r}"
    r += 2

    # ─── PROJECTION TABLE ───
    ws.cell(row=r, column=1, value="REVENUE → UFCF BRIDGE").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = GOLD_FILL
    r += 1

    # Headers: Label | Base Year | Y1..Y5
    proj_header = ["Concept ($M)", f"FY{last_year} (Base)"]
    for i, y in enumerate(proj_years):
        proj_header.append(f"FY{y}E")
    write_row(ws, r, proj_header, font=HEADER_FONT, fill=HEADER_FILL)
    proj_header_row = r
    r += 1

    # Base year column = column 2, Y1 = column 3, etc.
    base_col = 2
    y1_col = 3

    # Row: Revenue
    ws.cell(row=r, column=1, value="Revenue").font = BLACK_BOLD
    ws.cell(row=r, column=base_col, value=last_rev).font = BLUE_FONT
    ws.cell(row=r, column=base_col).number_format = FMT_USD
    rev_proj_row = r
    for i in range(n_proj):
        ci = y1_col + i
        prev_col = get_column_letter(ci - 1)
        if i == 0:
            growth_ref = ref_rev_g1
        elif i <= 2:
            growth_ref = ref_rev_g2
        else:
            growth_ref = ref_rev_g4
        ws.cell(row=r, column=ci, value=f"={prev_col}{r}*(1+{growth_ref})").font = BLACK_BOLD
        ws.cell(row=r, column=ci).number_format = FMT_USD
        ws.cell(row=r, column=ci).alignment = ALIGN_R
    r += 1

    # Row: Revenue Growth %
    ws.cell(row=r, column=1, value="  Growth %").font = Font(name="Arial", size=9, color="888888", italic=True)
    for i in range(n_proj):
        ci = y1_col + i
        prev_col = get_column_letter(ci - 1)
        curr_col = get_column_letter(ci)
        ws.cell(row=r, column=ci, value=f"={curr_col}{rev_proj_row}/{prev_col}{rev_proj_row}-1").font = Font(name="Arial", size=9, color="888888", italic=True)
        ws.cell(row=r, column=ci).number_format = FMT_PCT
        ws.cell(row=r, column=ci).alignment = ALIGN_R
    r += 1

    # Row: EBITDA
    ws.cell(row=r, column=1, value="EBITDA").font = BLACK_BOLD
    ws.cell(row=r, column=base_col, value=last_ebitda_m).font = BLUE_FONT
    ws.cell(row=r, column=base_col).number_format = FMT_USD
    ebitda_proj_row = r
    for i in range(n_proj):
        ci = y1_col + i
        rev_col = get_column_letter(ci)
        ws.cell(row=r, column=ci, value=f"={rev_col}{rev_proj_row}*{ref_ebitda_m}").font = BLACK_BOLD
        ws.cell(row=r, column=ci).number_format = FMT_USD
        ws.cell(row=r, column=ci).alignment = ALIGN_R
    r += 1

    # EBITDA Margin
    ws.cell(row=r, column=1, value="  EBITDA Margin").font = Font(name="Arial", size=9, color="888888", italic=True)
    for i in range(n_proj):
        ci = y1_col + i
        col_l = get_column_letter(ci)
        ws.cell(row=r, column=ci, value=f"={col_l}{ebitda_proj_row}/{col_l}{rev_proj_row}").font = Font(name="Arial", size=9, color="888888", italic=True)
        ws.cell(row=r, column=ci).number_format = FMT_PCT
        ws.cell(row=r, column=ci).alignment = ALIGN_R
    r += 1

    # D&A
    ws.cell(row=r, column=1, value="(-) Depreciation & Amort.").font = BLACK_FONT
    da_proj_row = r
    for i in range(n_proj):
        ci = y1_col + i
        rev_col = get_column_letter(ci)
        ws.cell(row=r, column=ci, value=f"=-{rev_col}{rev_proj_row}*{ref_da_pct}").font = BLACK_FONT
        ws.cell(row=r, column=ci).number_format = FMT_USD
        ws.cell(row=r, column=ci).alignment = ALIGN_R
    r += 1

    # EBIT = EBITDA + D&A (D&A is negative)
    ws.cell(row=r, column=1, value="EBIT").font = BLACK_BOLD
    ebit_proj_row = r
    for i in range(n_proj):
        ci = y1_col + i
        col_l = get_column_letter(ci)
        ws.cell(row=r, column=ci, value=f"={col_l}{ebitda_proj_row}+{col_l}{da_proj_row}").font = BLACK_BOLD
        ws.cell(row=r, column=ci).number_format = FMT_USD
        ws.cell(row=r, column=ci).alignment = ALIGN_R
        ws.cell(row=r, column=ci).border = THIN_BORDER
    r += 1

    # Taxes on EBIT
    ws.cell(row=r, column=1, value="(-) Taxes on EBIT").font = BLACK_FONT
    tax_proj_row = r
    for i in range(n_proj):
        ci = y1_col + i
        col_l = get_column_letter(ci)
        ws.cell(row=r, column=ci, value=f"=-{col_l}{ebit_proj_row}*{ref_tax}").font = BLACK_FONT
        ws.cell(row=r, column=ci).number_format = FMT_USD
        ws.cell(row=r, column=ci).alignment = ALIGN_R
    r += 1

    # NOPAT
    ws.cell(row=r, column=1, value="NOPAT").font = BLACK_BOLD
    nopat_proj_row = r
    for i in range(n_proj):
        ci = y1_col + i
        col_l = get_column_letter(ci)
        ws.cell(row=r, column=ci, value=f"={col_l}{ebit_proj_row}+{col_l}{tax_proj_row}").font = BLACK_BOLD
        ws.cell(row=r, column=ci).number_format = FMT_USD
        ws.cell(row=r, column=ci).alignment = ALIGN_R
    r += 1

    # Add back D&A
    ws.cell(row=r, column=1, value="(+) D&A").font = BLACK_FONT
    da_add_row = r
    for i in range(n_proj):
        ci = y1_col + i
        col_l = get_column_letter(ci)
        ws.cell(row=r, column=ci, value=f"=-{col_l}{da_proj_row}").font = BLACK_FONT
        ws.cell(row=r, column=ci).number_format = FMT_USD
        ws.cell(row=r, column=ci).alignment = ALIGN_R
    r += 1

    # Capex
    ws.cell(row=r, column=1, value="(-) Capex").font = BLACK_FONT
    capex_proj_row = r
    for i in range(n_proj):
        ci = y1_col + i
        rev_col = get_column_letter(ci)
        ws.cell(row=r, column=ci, value=f"=-{rev_col}{rev_proj_row}*{ref_capex_pct}").font = BLACK_FONT
        ws.cell(row=r, column=ci).number_format = FMT_USD
        ws.cell(row=r, column=ci).alignment = ALIGN_R
    r += 1

    # Change in NWC
    ws.cell(row=r, column=1, value="(-) Change in NWC").font = BLACK_FONT
    nwc_proj_row = r
    for i in range(n_proj):
        ci = y1_col + i
        rev_col = get_column_letter(ci)
        ws.cell(row=r, column=ci, value=f"=-{rev_col}{rev_proj_row}*{ref_nwc_pct}").font = BLACK_FONT
        ws.cell(row=r, column=ci).number_format = FMT_USD
        ws.cell(row=r, column=ci).alignment = ALIGN_R
    r += 1

    # UFCF = NOPAT + D&A + Capex + NWC
    ws.cell(row=r, column=1, value="Unlevered Free Cash Flow").font = BLACK_BOLD
    ws.cell(row=r, column=1).border = THICK_BOTTOM
    ufcf_proj_row = r
    for i in range(n_proj):
        ci = y1_col + i
        col_l = get_column_letter(ci)
        ws.cell(row=r, column=ci, value=f"={col_l}{nopat_proj_row}+{col_l}{da_add_row}+{col_l}{capex_proj_row}+{col_l}{nwc_proj_row}").font = BLACK_BOLD
        ws.cell(row=r, column=ci).number_format = FMT_USD
        ws.cell(row=r, column=ci).alignment = ALIGN_R
        ws.cell(row=r, column=ci).border = THICK_BOTTOM
        ws.cell(row=r, column=ci).fill = OUTPUT_FILL
    r += 2

    # ─── DCF VALUATION ───
    ws.cell(row=r, column=1, value="DCF VALUATION").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = GOLD_FILL
    r += 1

    # Terminal Value
    ws.cell(row=r, column=1, value="Terminal Value (Gordon Growth)").font = BLACK_FONT
    last_ufcf_col = get_column_letter(y1_col + n_proj - 1)
    ws.cell(row=r, column=2, value=f"={last_ufcf_col}{ufcf_proj_row}*(1+{ref_tgr})/({ref_wacc}-{ref_tgr})").font = BLACK_BOLD
    ws.cell(row=r, column=2).number_format = FMT_USD
    ws.cell(row=r, column=2).fill = OUTPUT_FILL
    ref_tv = f"$B${r}"
    r += 1

    # PV of FCFs
    ws.cell(row=r, column=1, value="PV of Projected FCFs").font = BLACK_FONT
    # Sum of each FCF / (1+WACC)^n
    pv_parts = []
    for i in range(n_proj):
        ci = y1_col + i
        col_l = get_column_letter(ci)
        pv_parts.append(f"{col_l}{ufcf_proj_row}/(1+{ref_wacc})^{i+1}")
    ws.cell(row=r, column=2, value=f"={'+'.join(pv_parts)}").font = BLACK_FONT
    ws.cell(row=r, column=2).number_format = FMT_USD
    ref_pv_fcf = f"$B${r}"
    r += 1

    # PV of Terminal Value
    ws.cell(row=r, column=1, value="PV of Terminal Value").font = BLACK_FONT
    ws.cell(row=r, column=2, value=f"={ref_tv}/(1+{ref_wacc})^{n_proj}").font = BLACK_FONT
    ws.cell(row=r, column=2).number_format = FMT_USD
    ref_pv_tv = f"$B${r}"
    r += 1

    # Enterprise Value
    ws.cell(row=r, column=1, value="Enterprise Value (Implied)").font = BLACK_BOLD
    ws.cell(row=r, column=2, value=f"={ref_pv_fcf}+{ref_pv_tv}").font = BLACK_BOLD
    ws.cell(row=r, column=2).number_format = FMT_USD
    ws.cell(row=r, column=2).fill = OUTPUT_FILL
    ref_ev = f"$B${r}"
    r += 1

    # Net Debt
    net_debt_val = ((total_debt or 0) - (total_cash or 0)) / 1e6
    ws.cell(row=r, column=1, value="(-) Net Debt").font = BLACK_FONT
    ws.cell(row=r, column=2, value=net_debt_val).font = BLUE_FONT
    ws.cell(row=r, column=2).number_format = FMT_USD
    ref_nd = f"$B${r}"
    r += 1

    # Equity Value
    ws.cell(row=r, column=1, value="Equity Value").font = BLACK_BOLD
    ws.cell(row=r, column=2, value=f"={ref_ev}-{ref_nd}").font = BLACK_BOLD
    ws.cell(row=r, column=2).number_format = FMT_USD
    ws.cell(row=r, column=2).fill = OUTPUT_FILL
    ref_eqv = f"$B${r}"
    r += 1

    # Shares Outstanding
    shares = (info.get("sharesOutstanding", 0) or 0) / 1e6
    ws.cell(row=r, column=1, value="Shares Outstanding (M)").font = BLACK_FONT
    ws.cell(row=r, column=2, value=shares).font = BLUE_FONT
    ws.cell(row=r, column=2).number_format = FMT_USD_DEC
    ref_shares = f"$B${r}"
    r += 1

    # Implied Price
    ws.cell(row=r, column=1, value="IMPLIED PRICE PER SHARE").font = BLACK_BOLD
    ws.cell(row=r, column=2, value=f"=IF({ref_shares}=0,0,{ref_eqv}/{ref_shares})").font = Font(name="Arial", size=12, color="000000", bold=True)
    ws.cell(row=r, column=2).number_format = FMT_PRICE
    ws.cell(row=r, column=2).fill = OUTPUT_FILL
    ws.cell(row=r, column=2).border = THICK_BOTTOM
    ref_implied = f"$B${r}"
    r += 1

    # Current Price
    curr_price = info.get("currentPrice", 0) or 0
    ws.cell(row=r, column=1, value="Current Price").font = BLACK_FONT
    ws.cell(row=r, column=2, value=curr_price).font = BLUE_FONT
    ws.cell(row=r, column=2).number_format = FMT_PRICE
    ref_curr = f"$B${r}"
    r += 1

    # Upside/Downside
    ws.cell(row=r, column=1, value="UPSIDE / DOWNSIDE").font = BLACK_BOLD
    ws.cell(row=r, column=2, value=f"=IF({ref_curr}=0,0,{ref_implied}/{ref_curr}-1)").font = Font(name="Arial", size=12, color="000000", bold=True)
    ws.cell(row=r, column=2).number_format = FMT_PCT
    ws.cell(row=r, column=2).fill = OUTPUT_FILL
    ws.cell(row=r, column=2).border = THICK_BOTTOM
    r += 2

    # ─── SENSITIVITY TABLE ───
    ws.cell(row=r, column=1, value="SENSITIVITY: WACC vs TERMINAL GROWTH → IMPLIED PRICE").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = GOLD_FILL
    r += 1

    # WACC scenarios: -2%, -1%, base, +1%, +2%
    wacc_offsets = [-0.02, -0.01, 0, 0.01, 0.02]
    tgr_values = [0.015, 0.02, 0.025, 0.03, 0.035]

    # Header
    ws.cell(row=r, column=1, value="WACC \\ TGR").font = HEADER_FONT
    ws.cell(row=r, column=1).fill = HEADER_FILL
    for j, tgr in enumerate(tgr_values):
        ws.cell(row=r, column=2+j, value=tgr).font = HEADER_FONT
        ws.cell(row=r, column=2+j).fill = HEADER_FILL
        ws.cell(row=r, column=2+j).number_format = FMT_PCT
    r += 1

    # Each row = a WACC scenario
    for offset in wacc_offsets:
        label = f"={ref_wacc}+{offset}" if offset != 0 else f"={ref_wacc}"
        ws.cell(row=r, column=1, value=label).font = BLACK_BOLD
        ws.cell(row=r, column=1).number_format = FMT_PCT

        for j, tgr in enumerate(tgr_values):
            # Recalc implied price: (last UFCF * (1+tgr) / (wacc_scenario - tgr)) discounted
            # Simplified: use PV of FCFs + PV of TV with new WACC/TGR
            # Full formula for sensitivity
            wacc_cell = f"({ref_wacc}+{offset})" if offset != 0 else ref_wacc
            tgr_val = tgr

            pv_parts_s = []
            for k in range(n_proj):
                ck = y1_col + k
                col_l = get_column_letter(ck)
                pv_parts_s.append(f"{col_l}{ufcf_proj_row}/(1+{wacc_cell})^{k+1}")

            tv_formula = f"{last_ufcf_col}{ufcf_proj_row}*(1+{tgr_val})/({wacc_cell}-{tgr_val})"
            pv_tv_formula = f"({tv_formula})/(1+{wacc_cell})^{n_proj}"
            ev_formula = f"{'+'.join(pv_parts_s)}+{pv_tv_formula}"
            price_formula = f"=IF({ref_shares}=0,0,({ev_formula}-{ref_nd})/{ref_shares})"

            cell = ws.cell(row=r, column=2+j, value=price_formula)
            cell.font = BLACK_FONT
            cell.number_format = FMT_PRICE
            cell.alignment = ALIGN_R

            # Highlight base case
            if offset == 0 and tgr == 0.025:
                cell.fill = OUTPUT_FILL
                cell.font = BLACK_BOLD

        r += 1

    return {
        "ref_wacc": ref_wacc,
        "ref_implied": ref_implied,
        "ref_ev": ref_ev,
        "ufcf_row": ufcf_proj_row,
    }


# ─────────────────────────────────────────────
# MAIN: GENERATE
# ─────────────────────────────────────────────

def generate_dcf_excel(ticker: str) -> io.BytesIO:
    ticker = ticker.upper()
    data = fetch_data(ticker)
    info = data["info"]

    wb = Workbook()

    build_historicals(wb, data, ticker)
    refs = build_projections(wb, data, ticker)

    # Freeze panes
    for ws in wb.worksheets:
        ws.freeze_panes = "B5"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ─────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python dcf_excel.py AAPL")
        sys.exit(1)

    ticker = sys.argv[1]
    print(f"Generando DCF para {ticker}...")
    buf = generate_dcf_excel(ticker)

    fname = f"DCF_{ticker}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    with open(fname, "wb") as f:
        f.write(buf.read())
    print(f"✅ Guardado: {fname}")
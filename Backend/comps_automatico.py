"""
📊 M&A COMPS AUTOMATICO v4
- Valores calculados en Python (no fórmulas Excel para márgenes/múltiplos)
- MEDIAN / AVERAGE simples para stats
- Implied Valuation con fórmulas Excel referenciando medianas
- Compatible con cualquier versión de Excel, cualquier idioma
"""

import yfinance as yf
import pandas as pd
import pathlib
import json
import time

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
import tempfile
# ─────────────────────────────────────────────
# 1. CONFIGURACIÓN DEL DEAL
# ─────────────────────────────────────────────

DEAL_CONFIG = {
    "empresa_target": "Swiss Medical",
    "sector":         "Health Insurance",
    "revenue_target": 1_000,
    "rango_min_pct":  0.3,
    "rango_max_pct":  3.0,
    "fecha":          datetime.now().strftime("%d/%m/%Y"),
}

# ─────────────────────────────────────────────
# 2. UNIVERSE DE EMPRESAS POR SECTOR
# ─────────────────────────────────────────────

def load_empresas():
    posibles_rutas = [
        "FrontEnd/Data/empresas.json",
        "Data/empresas.json",
        "../FrontEnd/Data/empresas.json"
    ]
    for ruta in posibles_rutas:
        path = pathlib.Path(ruta)
        if path.exists():
            return json.loads(path.read_text(encoding="utf-8"))
    return []

def get_universe_by_sector(sector):
    empresas = load_empresas()
    return [e["ticker"] for e in empresas if e.get("sector") == sector]

# ─────────────────────────────────────────────
# 3. FETCH DE DATOS — COMPS BÁSICOS
# ─────────────────────────────────────────────

def get_financials(ticker: str) -> dict | None:
    try:
        info = yf.Ticker(ticker).info
        if not info or info.get("quoteType") not in ["EQUITY"]:
            return None
        def mm(val):
            return round(val / 1_000_000, 1) if val else None
        desc = info.get("longBusinessSummary", "")
        if desc and len(desc) > 220:
            desc = desc[:217] + "..."
        return {
            "Ticker": ticker, "Empresa": info.get("shortName", ticker),
            "País": info.get("country", "N/A"), "Sector": info.get("sector", "N/A"),
            "Industria": info.get("industry", "N/A"), "Descripción": desc,
            "Revenue ($mm)": mm(info.get("totalRevenue")), "EBITDA ($mm)": mm(info.get("ebitda")),
            "Net Inc ($mm)": mm(info.get("netIncomeToCommon")), "Gross ($mm)": mm(info.get("grossProfits")),
            "Deuda ($mm)": mm(info.get("totalDebt")), "Cash ($mm)": mm(info.get("totalCash")),
            "Mkt Cap ($mm)": mm(info.get("marketCap")), "EV ($mm)": mm(info.get("enterpriseValue")),
            "P/E": round(info.get("trailingPE"), 1) if info.get("trailingPE") else None,
            "Rev Growth %": round(info.get("revenueGrowth") * 100, 1) if info.get("revenueGrowth") else None,
            "Empleados": info.get("fullTimeEmployees"),
        }
    except Exception:
        return None

# ─────────────────────────────────────────────
# 4. FETCH DE DATOS — DCF + WACC INPUTS
# ─────────────────────────────────────────────

def get_dcf_inputs(ticker: str) -> dict:
    try:
        t = yf.Ticker(ticker)
        info = t.info or {}
        def mm(val): return round(val / 1_000_000, 1) if val else None
        def pct(val): return round(val * 100, 2) if val else None

        revenue, ebitda, ebit = None, None, None
        try:
            q = t.quarterly_financials
            if q is not None and not q.empty:
                if "Total Revenue" in q.index:
                    revenue = round(q.loc["Total Revenue"].iloc[:4].sum() / 1_000_000, 1)
                if "EBITDA" in q.index:
                    ebitda = round(q.loc["EBITDA"].iloc[:4].sum() / 1_000_000, 1)
                if "EBIT" in q.index:
                    ebit = round(q.loc["EBIT"].iloc[:4].sum() / 1_000_000, 1)
        except Exception:
            pass

        if revenue is None: revenue = mm(info.get("totalRevenue"))
        if ebitda is None: ebitda = mm(info.get("ebitda"))
        if ebit is None and info.get("ebit"): ebit = mm(info.get("ebit"))

        net_income = mm(info.get("netIncomeToCommon"))
        tax_rate = pct(info.get("effectiveTaxRate"))
        interest_exp = mm(info.get("totalInterestExpense")) if info.get("totalInterestExpense") else None
        total_debt = mm(info.get("totalDebt"))
        cash = mm(info.get("totalCash"))
        net_debt = round(total_debt - cash, 1) if total_debt and cash else None
        mkt_cap = mm(info.get("marketCap"))
        ev = round(mkt_cap + total_debt - cash, 1) if mkt_cap and total_debt and cash is not None else mm(info.get("enterpriseValue"))
        capex = mm(info.get("capitalExpenditures"))
        fcf = mm(info.get("freeCashflow"))
        op_cf = mm(info.get("operatingCashflow"))
        da = round(ebitda - ebit, 1) if ebitda and ebit else None
        beta = round(info.get("beta"), 2) if info.get("beta") else None
        shares = mm(info.get("sharesOutstanding"))
        price = info.get("currentPrice") or info.get("regularMarketPrice")
        cost_of_debt = round(abs(interest_exp) / total_debt * 100, 2) if interest_exp and total_debt and total_debt > 0 else None
        debt_weight, equity_weight = None, None
        if mkt_cap and total_debt:
            total_capital = mkt_cap + total_debt
            equity_weight = round(mkt_cap / total_capital * 100, 1)
            debt_weight = round(total_debt / total_capital * 100, 1)
        net_debt_ebitda = round(net_debt / ebitda, 1) if net_debt and ebitda and ebitda > 0 else None

        return {
            "ticker": ticker, "empresa": info.get("shortName", ticker),
            "sector": info.get("sector", "N/A"), "pais": info.get("country", "N/A"),
            "revenue_mm": revenue, "ebitda_mm": ebitda, "ebit_mm": ebit,
            "net_income_mm": net_income, "tax_rate_pct": tax_rate, "interest_exp_mm": interest_exp,
            "ebitda_margin_pct": round(ebitda / revenue * 100, 1) if ebitda and revenue else None,
            "ebit_margin_pct": round(ebit / revenue * 100, 1) if ebit and revenue else None,
            "net_margin_pct": round(net_income / revenue * 100, 1) if net_income and revenue else None,
            "total_debt_mm": total_debt, "cash_mm": cash, "net_debt_mm": net_debt,
            "mkt_cap_mm": mkt_cap, "ev_mm": ev,
            "capex_mm": capex, "fcf_mm": fcf, "op_cashflow_mm": op_cf, "da_mm": da,
            "capex_pct_rev": round(abs(capex) / revenue * 100, 1) if capex and revenue else None,
            "fcf_margin_pct": round(fcf / revenue * 100, 1) if fcf and revenue else None,
            "beta": beta, "shares_mm": shares, "price": price,
            "cost_of_debt_pct": cost_of_debt, "equity_weight_pct": equity_weight,
            "debt_weight_pct": debt_weight, "net_debt_ebitda": net_debt_ebitda,
            "ev_revenue": round(ev / revenue, 1) if ev and revenue else None,
            "ev_ebitda": round(ev / ebitda, 1) if ev and ebitda else None,
            "pe": round(info.get("trailingPE"), 1) if info.get("trailingPE") else None,
            "pb": round(info.get("priceToBook"), 2) if info.get("priceToBook") else None,
            "rev_growth_pct": round(info.get("revenueGrowth") * 100, 1) if info.get("revenueGrowth") else None,
        }
    except Exception as e:
        return {"ticker": ticker, "error": str(e)}


def calcular_wacc(dcf_data: dict, risk_free_rate: float = 4.5, equity_risk_premium: float = 5.5) -> dict:
    beta = dcf_data.get("beta")
    cost_of_debt = dcf_data.get("cost_of_debt_pct")
    tax_rate = dcf_data.get("tax_rate_pct")
    equity_weight = dcf_data.get("equity_weight_pct")
    debt_weight = dcf_data.get("debt_weight_pct")

    result = {
        "risk_free_rate_pct": risk_free_rate, "equity_risk_premium_pct": equity_risk_premium,
        "beta": beta, "cost_of_equity_pct": None, "cost_of_debt_pct": cost_of_debt,
        "tax_rate_pct": tax_rate, "after_tax_cost_debt_pct": None,
        "equity_weight_pct": equity_weight, "debt_weight_pct": debt_weight,
        "wacc_pct": None, "nota": "",
    }
    if beta:
        cost_of_equity = risk_free_rate + beta * equity_risk_premium
        result["cost_of_equity_pct"] = round(cost_of_equity, 2)
    if cost_of_debt and tax_rate:
        after_tax_cod = cost_of_debt * (1 - tax_rate / 100)
        result["after_tax_cost_debt_pct"] = round(after_tax_cod, 2)
    if (result["cost_of_equity_pct"] and result["after_tax_cost_debt_pct"]
            and equity_weight and debt_weight):
        wacc = (result["cost_of_equity_pct"] * equity_weight / 100 +
                result["after_tax_cost_debt_pct"] * debt_weight / 100)
        result["wacc_pct"] = round(wacc, 2)
    elif result["cost_of_equity_pct"] and not cost_of_debt:
        result["wacc_pct"] = result["cost_of_equity_pct"]
        result["nota"] = "Cost of debt no disponible; WACC estimado como cost of equity"
    return result


# ─────────────────────────────────────────────
# 5. ESTILOS
# ─────────────────────────────────────────────

NAVY = "00205B"
DELOITTE = "86BC25"
LBLUE = "DCE6F1"
GRAY = "F5F5F5"
WHITE = "FFFFFF"
YELLOW = "FFFF00"

def hdr_font():   return Font(name="Arial", bold=True, color="FFFFFF", size=9)
def dat_font():   return Font(name="Arial", size=9)
def bold_font():  return Font(name="Arial", bold=True, size=9)
def blue_font():  return Font(name="Arial", size=9, color="0000FF")
def title_font(): return Font(name="Arial", bold=True, size=13, color=DELOITTE)
def sub_font():   return Font(name="Arial", size=9, italic=True, color="666666")
def hdr_fill():   return PatternFill("solid", start_color=DELOITTE)
def sum_fill():   return PatternFill("solid", start_color=LBLUE)
def alt_fill(i):  return PatternFill("solid", start_color=GRAY if i % 2 == 0 else WHITE)

thin = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
AL = Alignment(horizontal="left", vertical="center")
AR = Alignment(horizontal="right", vertical="center")
AC = Alignment(horizontal="center", vertical="center")
AW = Alignment(horizontal="left", vertical="top", wrap_text=True)

# ─────────────────────────────────────────────
# 6. COLUMNAS Y HELPERS
# ─────────────────────────────────────────────

COLS = [
    ("Ticker",        9,  "txt"),
    ("Empresa",       30, "txt"),
    ("País",          14, "txt"),
    ("Sector",        18, "txt"),
    ("Industria",     24, "txt"),
    ("Descripción",   60, "wrap"),
    ("Revenue ($mm)", 15, "num"),
    ("EBITDA ($mm)",  14, "num"),
    ("Net Inc ($mm)", 14, "num"),
    ("Gross ($mm)",   13, "num"),
    ("Deuda ($mm)",   13, "num"),
    ("Cash ($mm)",    13, "num"),
    ("Mkt Cap ($mm)", 15, "num"),
    ("EV ($mm)",      14, "num"),
    ("P/E",           9,  "mult"),
    ("Rev Growth %",  13, "pct"),
    ("Empleados",     13, "int"),
    ("EBITDA Mg%",    13, "val_pct"),
    ("Net Mg%",       12, "val_pct"),
    ("Gross Mg%",     13, "val_pct"),
    ("EV/Revenue",    13, "val_mult"),
    ("EV/EBITDA",     13, "val_mult"),
]

COLS_VISIBLE = [c for c in COLS if len(c) == 3]


def write_header_row(ws, row, cols):
    for ci, (name, width, _) in enumerate(cols, 1):
        cell = ws.cell(row=row, column=ci, value=name)
        cell.font = hdr_font()
        cell.fill = hdr_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[row].height = 30


def write_data_row(ws, er, row_data, cols, col_idx_map):
    fill = alt_fill(er)

    for ci, (key, width, fmt) in enumerate(cols, 1):
        val = row_data.get(key)
        cell = ws.cell(row=er, column=ci, value=val)
        cell.fill = fill
        cell.border = thin

        if fmt == "wrap":
            cell.font = dat_font(); cell.alignment = AW
        elif fmt == "num":
            cell.font = dat_font(); cell.alignment = AR; cell.number_format = '#,##0.0'
        elif fmt == "mult":
            cell.font = dat_font(); cell.alignment = AR; cell.number_format = '0.0"x"'
        elif fmt == "pct":
            cell.font = dat_font(); cell.alignment = AR; cell.number_format = '0.0"%"'
        elif fmt == "int":
            cell.font = dat_font(); cell.alignment = AR; cell.number_format = '#,##0'
        elif fmt in ("val_pct", "val_mult"):
            cell.value = None  # se llena abajo
        else:
            cell.font = dat_font(); cell.alignment = AL

    # ── VALORES CALCULADOS EN PYTHON (no fórmulas Excel) ──
    rev_val = row_data.get("Revenue ($mm)")
    ebitda_val = row_data.get("EBITDA ($mm)")
    net_val = row_data.get("Net Inc ($mm)")
    ev_val = row_data.get("EV ($mm)")
    gross_val = row_data.get("Gross ($mm)")

    calc_map = {
        "EBITDA Mg%": (round(ebitda_val / rev_val * 100, 1) if ebitda_val and rev_val and rev_val > 0 else None, '0.0"%"'),
        "Net Mg%":    (round(net_val / rev_val * 100, 1) if net_val and rev_val and rev_val > 0 else None, '0.0"%"'),
        "Gross Mg%":  (round(gross_val / rev_val * 100, 1) if gross_val and rev_val and rev_val > 0 else None, '0.0"%"'),
        "EV/Revenue": (round(ev_val / rev_val, 1) if ev_val and rev_val and rev_val > 0 else None, '0.0"x"'),
        "EV/EBITDA":  (round(ev_val / ebitda_val, 1) if ev_val and ebitda_val and ebitda_val > 0 else None, '0.0"x"'),
    }

    for key, (value, fmt_str) in calc_map.items():
        if key not in col_idx_map:
            continue
        cell = ws.cell(row=er, column=col_idx_map[key], value=value)
        cell.fill = fill
        cell.border = thin
        cell.font = dat_font()
        cell.alignment = AR
        cell.number_format = fmt_str


# ─────────────────────────────────────────────
# NUM_STAT_COLS — columnas con Mediana/Promedio
# ─────────────────────────────────────────────

NUM_STAT_COLS = [
    ("Revenue ($mm)", '#,##0.0'), ("EBITDA ($mm)", '#,##0.0'),
    ("Net Inc ($mm)", '#,##0.0'), ("EV ($mm)", '#,##0.0'),
    ("Mkt Cap ($mm)", '#,##0.0'), ("P/E", '0.0"x"'),
    ("Rev Growth %", '0.0"%"'),
    ("EBITDA Mg%", '0.0"%"'), ("Net Mg%", '0.0"%"'), ("Gross Mg%", '0.0"%"'),
    ("EV/Revenue", '0.0"x"'), ("EV/EBITDA", '0.0"x"'),
]


def write_stats_rows(ws, data_start, last_row, col_idx_map):
    """Escribe Mediana y Promedio con =MEDIAN() y =AVERAGE() simples."""
    med_row = last_row + 2
    avg_row = last_row + 3

    for offset, label in enumerate(["Mediana", "Promedio"], 1):
        sr = last_row + 1 + offset
        ws.cell(row=sr, column=1, value=label).font = bold_font()
        ws.cell(row=sr, column=1).fill = sum_fill()
        ws.cell(row=sr, column=1).border = thin
        fn = "MEDIAN" if label == "Mediana" else "AVERAGE"
        for key, fmt_str in NUM_STAT_COLS:
            if key not in col_idx_map:
                continue
            ci = col_idx_map[key]
            cl = get_column_letter(ci)
            cell = ws.cell(row=sr, column=ci, value=f"={fn}({cl}{data_start}:{cl}{last_row})")
            cell.fill = sum_fill()
            cell.border = thin
            cell.font = bold_font()
            cell.alignment = AR
            cell.number_format = fmt_str

    return med_row, avg_row


# ─────────────────────────────────────────────
# 7. GENERAR EXCEL — COMPS (standalone)
# ─────────────────────────────────────────────

def generar_excel(df: pd.DataFrame):
    wb = Workbook()
    df_sorted = df.sort_values("Revenue ($mm)", ascending=False).reset_index(drop=True)
    col_idx_map = {c[0]: i for i, c in enumerate(COLS_VISIBLE, 1)}

    ws1 = wb.active
    ws1.title = "Universe Completo"
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = "A4"

    ncols = len(COLS_VISIBLE)
    ws1.merge_cells(f"A1:{get_column_letter(ncols)}1")
    ws1["A1"] = f"Comparable Companies — Universe | {DEAL_CONFIG['empresa_target']} | {DEAL_CONFIG['sector']}"
    ws1["A1"].font = title_font()
    ws1.row_dimensions[1].height = 26

    ws1.merge_cells(f"A2:{get_column_letter(ncols)}2")
    ws1["A2"] = f"Analista: {DEAL_CONFIG.get('analista', 'Analista')}  |  Fecha: {DEAL_CONFIG['fecha']}  |  Fuente: Yahoo Finance  |  {len(df_sorted)} empresas"
    ws1["A2"].font = sub_font()
    ws1.row_dimensions[2].height = 16

    write_header_row(ws1, 4, COLS_VISIBLE)

    DATA_START1 = 5
    for i, row_data in df_sorted.iterrows():
        er = DATA_START1 + i
        ws1.row_dimensions[er].height = 32 if row_data.get("Descripción") else 18
        write_data_row(ws1, er, row_data, COLS_VISIBLE, col_idx_map)

    last1 = DATA_START1 + len(df_sorted) - 1
    write_stats_rows(ws1, DATA_START1, last1, col_idx_map)

    # HOJA 2: FILTRADAS
    ws2 = wb.create_sheet("Comps Filtradas")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A8"

    ws2["A1"] = "PARÁMETROS DEL DEAL"
    ws2["A1"].font = Font(name="Arial", bold=True, size=11, color=DELOITTE)
    ws2.row_dimensions[1].height = 20

    param_rows = [
        (2, "Revenue Target ($mm)", DEAL_CONFIG["revenue_target"], '#,##0'),
        (3, "Rango mínimo (%)", DEAL_CONFIG["rango_min_pct"] * 100, '0"%"'),
        (4, "Rango máximo (%)", DEAL_CONFIG["rango_max_pct"] * 100, '0"%"'),
    ]
    for r_num, label, val, fmt_str in param_rows:
        ws2.cell(row=r_num, column=1, value=label).font = dat_font()
        cell = ws2.cell(row=r_num, column=2, value=val)
        cell.font = blue_font()
        cell.fill = PatternFill("solid", start_color=YELLOW)
        cell.number_format = fmt_str
        cell.border = thin

    ws2.merge_cells(f"A6:{get_column_letter(ncols)}6")
    ws2["A6"] = "Empresas filtradas por revenue range"
    ws2["A6"].font = sub_font()

    write_header_row(ws2, 7, COLS_VISIBLE)

    rev_min = DEAL_CONFIG["revenue_target"] * DEAL_CONFIG["rango_min_pct"]
    rev_max = DEAL_CONFIG["revenue_target"] * DEAL_CONFIG["rango_max_pct"]
    df_filt = df_sorted[
        (df_sorted["Revenue ($mm)"].notna()) &
        (df_sorted["Revenue ($mm)"] >= rev_min) &
        (df_sorted["Revenue ($mm)"] <= rev_max)
    ].reset_index(drop=True)

    DATA_START2 = 8
    for i, row_data in df_filt.iterrows():
        er = DATA_START2 + i
        ws2.row_dimensions[er].height = 42
        write_data_row(ws2, er, row_data, COLS_VISIBLE, col_idx_map)

    last2 = DATA_START2 + len(df_filt) - 1
    med_row, avg_row = write_stats_rows(ws2, DATA_START2, last2, col_idx_map)

    # IMPLIED VALUATION
    ev_rev_ci = col_idx_map.get("EV/Revenue")
    ev_ebt_ci = col_idx_map.get("EV/EBITDA")
    ebt_mg_ci = col_idx_map.get("EBITDA Mg%")

    if ev_rev_ci and ev_ebt_ci and ebt_mg_ci:
        ev_rev_ref = f"{get_column_letter(ev_rev_ci)}{med_row}"
        ev_ebt_ref = f"{get_column_letter(ev_ebt_ci)}{med_row}"
        ebt_mg_ref = f"{get_column_letter(ebt_mg_ci)}{med_row}"

        impl = avg_row + 3
        ws2.cell(row=impl, column=1, value="── IMPLIED VALUATION ──").font = Font(name="Arial", bold=True, size=10, color=DELOITTE)
        ws2.row_dimensions[impl].height = 20

        impl_rows = [
            ("Revenue Target ($mm)", "=$B$2", '#,##0'),
            ("EV Implied — EV/Revenue (mediana)", f"=$B$2*{ev_rev_ref}", '#,##0'),
            ("EV Implied — EV/EBITDA (mediana)", f"=$B$2*({ebt_mg_ref}/100)*{ev_ebt_ref}", '#,##0'),
        ]
        for j, (lbl, formula, fmt_str) in enumerate(impl_rows):
            r = impl + 1 + j
            ws2.cell(row=r, column=1, value=lbl).font = dat_font()
            cell = ws2.cell(row=r, column=2, value=formula)
            cell.font = bold_font()
            cell.fill = sum_fill()
            cell.border = thin
            cell.alignment = AR
            cell.number_format = fmt_str

    fecha_str = datetime.now().strftime("%Y%m%d")
    fname = f"Comps_{DEAL_CONFIG['empresa_target'].replace(' ', '_')}_{fecha_str}.xlsx"
    wb.save(fname)
    print(f"\n  ✅ Excel guardado: {fname}")
    return fname


def _generar_excel_buffer(df: pd.DataFrame, buffer, df_universe: pd.DataFrame = None) -> None:
    """Genera Excel con 2 hojas: Universe Completo + Comps Filtradas."""
    wb = Workbook()
    col_idx_map = {c[0]: i for i, c in enumerate(COLS_VISIBLE, 1)}
    ncols = len(COLS_VISIBLE)

    # ══════════════════════════════════════════════
    # HOJA 1: UNIVERSE COMPLETO
    # ══════════════════════════════════════════════
    if df_universe is not None and not df_universe.empty:
        df_uni = df_universe.sort_values("Revenue ($mm)", ascending=False).reset_index(drop=True)
    else:
        df_uni = df.sort_values("Revenue ($mm)", ascending=False).reset_index(drop=True)

    ws1 = wb.active
    ws1.title = "Industry Universe"
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = "A5"

    ws1.merge_cells(f"A1:{get_column_letter(ncols)}1")
    ws1["A1"] = f"Industry Universe — {DEAL_CONFIG['empresa_target']} | {DEAL_CONFIG.get('sector', '')}"
    ws1["A1"].font = title_font()
    ws1.row_dimensions[1].height = 26

    ws1.merge_cells(f"A2:{get_column_letter(ncols)}2")
    ws1["A2"] = f"Fecha: {DEAL_CONFIG['fecha']}  |  Fuente: Yahoo Finance (TTM)  |  {len(df_uni)} empresas  |  DealDesk"
    ws1["A2"].font = sub_font()
    ws1.row_dimensions[2].height = 16

    # Empty row 3
    ws1.row_dimensions[3].height = 6

    write_header_row(ws1, 4, COLS_VISIBLE)

    DATA_START_1 = 5
    for i, row_data in df_uni.iterrows():
        er = DATA_START_1 + i
        ws1.row_dimensions[er].height = 32 if row_data.get("Descripción") else 18
        write_data_row(ws1, er, row_data, COLS_VISIBLE, col_idx_map)

    last_1 = DATA_START_1 + len(df_uni) - 1
    write_stats_rows(ws1, DATA_START_1, last_1, col_idx_map)

    # ══════════════════════════════════════════════
    # HOJA 2: COMPS FILTRADAS + IMPLIED VALUATION
    # ══════════════════════════════════════════════
    df_sorted = df.sort_values("Revenue ($mm)", ascending=False).reset_index(drop=True)

    ws2 = wb.create_sheet("Comparable Companies")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A9"

    # ── PARÁMETROS DEL DEAL ──
    ws2["A1"] = "PARÁMETROS DEL DEAL"
    ws2["A1"].font = Font(name="Arial", bold=True, size=11, color=DELOITTE)
    ws2.row_dimensions[1].height = 20

    rev_target = DEAL_CONFIG.get("revenue_target", 0)
    param_rows = [
        (2, "Revenue Target ($mm)", rev_target, '#,##0'),
        (3, "Rango mínimo (%)", DEAL_CONFIG.get("rango_min_pct", 0.3) * 100, '0"%"'),
        (4, "Rango máximo (%)", DEAL_CONFIG.get("rango_max_pct", 3.0) * 100, '0"%"'),
    ]
    for r_num, label, val, fmt_str in param_rows:
        ws2.cell(row=r_num, column=1, value=label).font = dat_font()
        cell = ws2.cell(row=r_num, column=2, value=val)
        cell.font = blue_font()
        cell.fill = PatternFill("solid", start_color=YELLOW)
        cell.number_format = fmt_str
        cell.border = thin

    ws2.row_dimensions[5].height = 6  # spacer

    ws2.merge_cells(f"A6:{get_column_letter(ncols)}6")
    ws2["A6"] = f"Comparable Companies — {DEAL_CONFIG['empresa_target']} | {len(df_sorted)} comps filtradas"
    ws2["A6"].font = title_font()
    ws2.row_dimensions[6].height = 24

    ws2.merge_cells(f"A7:{get_column_letter(ncols)}7")
    ws2["A7"] = f"Fecha: {DEAL_CONFIG['fecha']}  |  Fuente: Yahoo Finance (TTM)  |  DealDesk"
    ws2["A7"].font = sub_font()
    ws2.row_dimensions[7].height = 16

    write_header_row(ws2, 8, COLS_VISIBLE)

    DATA_START_2 = 9
    for i, row_data in df_sorted.iterrows():
        er = DATA_START_2 + i
        ws2.row_dimensions[er].height = 32 if row_data.get("Descripción") else 18
        write_data_row(ws2, er, row_data, COLS_VISIBLE, col_idx_map)

    last_2 = DATA_START_2 + len(df_sorted) - 1
    med_row, avg_row = write_stats_rows(ws2, DATA_START_2, last_2, col_idx_map)

    # ── IMPLIED VALUATION ──
    ev_rev_ci = col_idx_map.get("EV/Revenue")
    ev_ebt_ci = col_idx_map.get("EV/EBITDA")
    ebt_mg_ci = col_idx_map.get("EBITDA Mg%")

    if ev_rev_ci and ev_ebt_ci and ebt_mg_ci:
        ev_rev_ref = f"{get_column_letter(ev_rev_ci)}{med_row}"
        ev_ebt_ref = f"{get_column_letter(ev_ebt_ci)}{med_row}"
        ebt_mg_ref = f"{get_column_letter(ebt_mg_ci)}{med_row}"

        impl = avg_row + 3
        ws2.cell(row=impl, column=1, value="── IMPLIED VALUATION ──").font = Font(name="Arial", bold=True, size=10, color=DELOITTE)
        ws2.row_dimensions[impl].height = 20

        impl_rows = [
            ("Revenue Target ($mm)", "=$B$2", '#,##0'),
            ("EV Implied — EV/Revenue (mediana)", f"=$B$2*{ev_rev_ref}" if rev_target else "", '#,##0'),
            ("EV Implied — EV/EBITDA (mediana)", f"=$B$2*({ebt_mg_ref}/100)*{ev_ebt_ref}" if rev_target else "", '#,##0'),
        ]
        for j, (lbl, formula, fmt_str) in enumerate(impl_rows):
            r = impl + 1 + j
            ws2.cell(row=r, column=1, value=lbl).font = dat_font()
            cell = ws2.cell(row=r, column=2, value=formula)
            cell.font = bold_font()
            cell.fill = sum_fill()
            cell.border = thin
            cell.alignment = AR
            cell.number_format = fmt_str

    # Hoja 2 activa por defecto (la que el analista necesita ver primero)
    wb.active = wb.sheetnames.index("Comparable Companies")
    # ══════════════════════════════════════════════
    # 📊 CHARTS (EXCEL REAL)
    # ══════════════════════════════════════════════

    try:
        from openpyxl.chart import ScatterChart, Reference, Series, BarChart

        ws3 = wb.create_sheet("Charts")

        # ─────────────────────────────
        # TABLA BASE (datos para charts)
        # ─────────────────────────────

        ws3["A1"] = "Ticker"
        ws3["B1"] = "Revenue"
        ws3["C1"] = "EV"
        ws3["D1"] = "EV/EBITDA"

        row_start = 2

        for i, row in df.iterrows():
            r = row_start + i

            ws3.cell(r, 1, row["Ticker"])
            ws3.cell(r, 2, row["Revenue ($mm)"])
            ws3.cell(r, 3, row["EV ($mm)"])
            ws3.cell(r, 4, row.get("EV/EBITDA"))

        last_row = row_start + len(df) - 1

        # ─────────────────────────────
        # SCATTER: EV vs Revenue
        # ─────────────────────────────

        scatter = ScatterChart()
        scatter.title = "EV vs Revenue"
        scatter.style = 13

        xvalues = Reference(ws3, min_col=2, min_row=2, max_row=last_row)
        yvalues = Reference(ws3, min_col=3, min_row=2, max_row=last_row)

        series = Series(yvalues, xvalues, title="Comps")
        scatter.series.append(series)

        scatter.x_axis.title = "Revenue ($mm)"
        scatter.y_axis.title = "Enterprise Value ($mm)"

        scatter.height = 10
        scatter.width = 18

        ws3.add_chart(scatter, "F2")

        # ─────────────────────────────
        # BAR: EV / EBITDA
        # ─────────────────────────────

        bar = BarChart()
        bar.title = "EV / EBITDA"
        bar.style = 10

        data = Reference(ws3, min_col=4, min_row=1, max_row=last_row)
        cats = Reference(ws3, min_col=1, min_row=2, max_row=last_row)

        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)

        bar.height = 10
        bar.width = 18

        ws3.add_chart(bar, "F20")

    except Exception as e:
        print("⚠️ Error generando charts:", e)
    wb.save(buffer)


# ─────────────────────────────────────────────
# 8. MAIN
# ─────────────────────────────────────────────

def run():
    print("\n" + "=" * 60)
    print(f"  📊 M&A COMPS v4 — {DEAL_CONFIG['empresa_target']}")
    print(f"  Sector: {DEAL_CONFIG['sector']}  |  Revenue target: ${DEAL_CONFIG['revenue_target']:,}mm")
    print("=" * 60)

    sector = DEAL_CONFIG['sector']
    tickers = get_universe_by_sector(sector)
    print(f"\n  Descargando datos de {len(tickers)} empresas...\n")

    resultados = []
    for i, ticker in enumerate(tickers):
        print(f"  [{i + 1:02d}/{len(tickers)}] {ticker:<12}", end=" ")
        data = get_financials(ticker)
        if data and data.get("Revenue ($mm)"):
            resultados.append(data)
            print(f"✅  ${data['Revenue ($mm)']:>10,.0f}mm  {data.get('Empresa', '')}")
        else:
            print("⚠️  sin datos")
        time.sleep(0.25)

    if not resultados:
        print("\n  ❌ Sin datos.")
        return

    df = pd.DataFrame(resultados)
    generar_excel(df)


if __name__ == "__main__":
    run()
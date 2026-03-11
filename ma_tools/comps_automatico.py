"""
📊 M&A COMPS AUTOMATICO v3
- Motor financiero universal: comps, DCF inputs, WACC, precedents
- 50 empresas comparables por sector
- Hoja 1: Universe completo con descripción
- Hoja 2: Filtradas por revenue (fórmulas Excel)
- Márgenes calculados con fórmulas, no hardcodeados
"""

import yfinance as yf
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import time

# ─────────────────────────────────────────────
# 1. CONFIGURACIÓN DEL DEAL
# ─────────────────────────────────────────────

DEAL_CONFIG = {
    "empresa_target": "Swiss Medical",
    "sector":         "Health Insurance",
    "revenue_target": 1_000,
    "rango_min_pct":  0.3,
    "rango_max_pct":  3.0,
    "analista":       "Francisco Pini",
    "fecha":          datetime.now().strftime("%d/%m/%Y"),
}

# ─────────────────────────────────────────────
# 2. UNIVERSE DE EMPRESAS POR SECTOR
# ─────────────────────────────────────────────

UNIVERSE = {
    "Health Insurance": [
        "UNH", "CVS", "CI", "ELV", "CNC", "HUM", "MOH", "OSCR", "CLOV", "ALHC",
        "MMC", "AON", "WTW", "ACGL", "THG", "HIG", "PRU", "MET", "AFL", "UNM",
        "HCA", "THC", "UHS", "CYH", "ENSG",
        "HQY", "SGRY", "NHC", "PINC", "DVA",
        "JNJ", "ABT", "MDT", "BMY", "AMGN",
        "IQV", "DGX", "LH", "MDRX", "DOCS",
        "HIMS", "ACAD", "NVCR", "RDDT", "NHI",
        "WELL", "VTR", "PEAK", "MPW", "SBRA",
    ],
    "Technology": [
        "AAPL", "MSFT", "GOOGL", "META", "AMZN", "NVDA", "CRM", "ORCL", "SAP", "IBM",
        "ADBE", "INTU", "NOW", "SNOW", "PLTR", "DDOG", "MDB", "TWLO", "ZS", "CRWD",
        "NET", "OKTA", "HUBS", "BILL", "SMAR", "APPF", "PCTY", "PAYC", "VEEV", "DOCU",
        "ZM", "WDAY", "CDNS", "ANSS", "PTC", "EPAM", "GLOB", "MELI", "SE", "GRAB",
        "BIDU", "JD", "PDD", "TCEHY", "NTES", "SHOP", "SPOT", "UBER", "LYFT", "DASH",
    ],
    "Financials": [
        "JPM", "BAC", "WFC", "GS", "MS", "C", "AXP", "BLK", "SCHW", "USB",
        "PNC", "TFC", "COF", "DFS", "SYF", "ALLY", "NDAQ", "ICE", "CME", "CBOE",
        "BX", "KKR", "APO", "CG", "ARES", "OWL", "IVZ", "BEN", "AMG", "TROW",
        "BK", "STT", "NTRS", "FNF", "FAF", "RJF", "SF", "LPLA", "MORN", "FDS",
        "MSCI", "SPGI", "MCO", "VRSK", "BR", "FIS", "FI", "GPN", "MA", "V",
    ],
    "Energy": [
        "XOM", "CVX", "COP", "EOG", "SLB", "MPC", "PSX", "VLO", "OXY", "DVN",
        "HAL", "BKR", "FANG", "PR", "CTRA", "APA", "MRO", "HES", "MTDR", "SM",
        "RIG", "VAL", "DO", "NE", "HP", "PTEN", "WES", "TRGP", "KMI", "OKE",
        "WMB", "ET", "EPD", "MMP", "MPLX", "PAA", "LNG", "RRC", "EQT", "AR",
        "CHK", "SWN", "GPOR", "CNX", "CDEV", "ESTE", "NOG", "VTLE", "SBOW", "PDCE",
    ],
    "Consumer": [
        "AMZN", "WMT", "COST", "TGT", "HD", "LOW", "MCD", "SBUX", "NKE", "LULU",
        "TJX", "ROST", "DG", "DLTR", "KR", "ACI", "SFM", "WINN", "GO", "CASY",
        "YUM", "QSR", "DPZ", "CMG", "DRI", "TXRH", "SHAK", "JACK", "WEN", "BJ",
        "EL", "PG", "CL", "KMB", "CHD", "SPB", "HRL", "SJM", "MKC", "GIS",
        "K", "CPB", "CAG", "MDLZ", "HSY", "MNST", "KO", "PEP", "STZ", "BUD",
    ],
    "Real Estate": [
        "PLD", "AMT", "CCI", "EQIX", "SPG", "O", "VICI", "WELL", "DLR", "PSA",
        "EXR", "AVB", "EQR", "MAA", "UDR", "CPT", "NNN", "WPC", "STAG", "COLD",
        "IRM", "SBAC", "SBA", "AMH", "INVH", "SUI", "ELS", "UE", "KIM", "REG",
        "BXP", "VNO", "SL", "HIW", "CUZ", "PDM", "OFC", "DEI", "PGRE", "ESRT",
        "HST", "RHP", "PK", "SHO", "APLE", "CLDT", "NCLH", "RLJ", "AHT", "BHR",
    ],
    "Industrials": [
        "GE", "HON", "MMM", "CAT", "DE", "EMR", "ETN", "PH", "ROK", "ITW",
        "GD", "LMT", "RTX", "NOC", "BA", "HII", "L3H", "LDOS", "SAIC", "CACI",
        "UPS", "FDX", "XPO", "JBHT", "CHRW", "EXPD", "GXO", "ODFL", "SAIA", "RXO",
        "WM", "RSG", "CWST", "CLH", "US", "SRCL", "ACCO", "CTAS", "ABM", "SCI",
        "AME", "DHR", "ROP", "VRTS", "IDEX", "IEX", "FTV", "GNRC", "AXON", "TDY",
    ],
}

# ─────────────────────────────────────────────
# 3. FETCH DE DATOS — COMPS BÁSICOS
# ─────────────────────────────────────────────

def get_financials(ticker: str) -> dict | None:
    """Datos para tabla de comps: múltiplos, márgenes, tamaño."""
    try:
        info = yf.Ticker(ticker).info
        if not info or info.get("quoteType") not in ["EQUITY"]:
            return None

        def mm(val):
            return round(val / 1_000_000, 1) if val else None

        desc = info.get("longBusinessSummary", "")
        if desc and len(desc) > 220:
            desc = desc[:217] + "..."

        return {
            "Ticker":        ticker,
            "Empresa":       info.get("shortName", ticker),
            "País":          info.get("country", "N/A"),
            "Sector":        info.get("sector", "N/A"),
            "Industria":     info.get("industry", "N/A"),
            "Descripción":   desc,
            "Revenue ($mm)": mm(info.get("totalRevenue")),
            "EBITDA ($mm)":  mm(info.get("ebitda")),
            "Net Inc ($mm)": mm(info.get("netIncomeToCommon")),
            "Gross ($mm)":   mm(info.get("grossProfits")),
            "Deuda ($mm)":   mm(info.get("totalDebt")),
            "Cash ($mm)":    mm(info.get("totalCash")),
            "Mkt Cap ($mm)": mm(info.get("marketCap")),
            "EV ($mm)":      mm(info.get("enterpriseValue")),
            "P/E":           round(info.get("trailingPE"), 1) if info.get("trailingPE") else None,
            "Rev Growth %":  round(info.get("revenueGrowth") * 100, 1) if info.get("revenueGrowth") else None,
            "Empleados":     info.get("fullTimeEmployees"),
        }
    except Exception:
        return None


# ─────────────────────────────────────────────
# 4. FETCH DE DATOS — DCF + WACC INPUTS
# ─────────────────────────────────────────────

def get_dcf_inputs(ticker: str) -> dict:
    """
    Trae todos los inputs necesarios para DCF y WACC:
    - Beta, cost of debt, tax rate
    - CapEx, D&A, FCF
    - Working capital
    - Shares outstanding, price
    """
    try:
        t    = yf.Ticker(ticker)
        info = t.info or {}

        def mm(val):
            return round(val / 1_000_000, 1) if val else None

        def pct(val):
            return round(val * 100, 2) if val else None

        # ── Income Statement inputs ──
        revenue      = mm(info.get("totalRevenue"))
        ebitda       = mm(info.get("ebitda"))
        ebit         = mm(info.get("ebit")) if info.get("ebit") else None
        net_income   = mm(info.get("netIncomeToCommon"))
        tax_rate     = pct(info.get("effectiveTaxRate"))
        interest_exp = mm(info.get("totalInterestExpense")) if info.get("totalInterestExpense") else None

        # ── Balance Sheet inputs ──
        total_debt   = mm(info.get("totalDebt"))
        cash         = mm(info.get("totalCash"))
        net_debt     = round(total_debt - cash, 1) if total_debt and cash else None
        mkt_cap      = mm(info.get("marketCap"))
        ev           = mm(info.get("enterpriseValue"))

        # ── Cash Flow inputs ──
        capex        = mm(info.get("capitalExpenditures"))
        fcf          = mm(info.get("freeCashflow"))
        op_cf        = mm(info.get("operatingCashflow"))

        # D&A: estimado como EBITDA - EBIT si disponibles
        da = None
        if ebitda and ebit:
            da = round(ebitda - ebit, 1)

        # ── WACC inputs ──
        beta         = round(info.get("beta"), 2) if info.get("beta") else None
        shares       = mm(info.get("sharesOutstanding"))
        price        = info.get("currentPrice") or info.get("regularMarketPrice")

        # Cost of Debt estimado: interest / total_debt
        cost_of_debt = None
        if interest_exp and total_debt and total_debt > 0:
            cost_of_debt = round(abs(interest_exp) / total_debt * 100, 2)

        # Weights para WACC
        debt_weight   = None
        equity_weight = None
        if mkt_cap and total_debt:
            total_capital  = mkt_cap + total_debt
            equity_weight  = round(mkt_cap / total_capital * 100, 1)
            debt_weight    = round(total_debt / total_capital * 100, 1)

        # Leverage
        net_debt_ebitda = round(net_debt / ebitda, 1) if net_debt and ebitda and ebitda > 0 else None

        return {
            # Identificación
            "ticker":          ticker,
            "empresa":         info.get("shortName", ticker),
            "sector":          info.get("sector", "N/A"),
            "pais":            info.get("country", "N/A"),

            # P&L
            "revenue_mm":      revenue,
            "ebitda_mm":       ebitda,
            "ebit_mm":         ebit,
            "net_income_mm":   net_income,
            "tax_rate_pct":    tax_rate,
            "interest_exp_mm": interest_exp,

            # Márgenes
            "ebitda_margin_pct": round(ebitda / revenue * 100, 1) if ebitda and revenue else None,
            "ebit_margin_pct":   round(ebit   / revenue * 100, 1) if ebit and revenue else None,
            "net_margin_pct":    round(net_income / revenue * 100, 1) if net_income and revenue else None,

            # Balance
            "total_debt_mm":   total_debt,
            "cash_mm":         cash,
            "net_debt_mm":     net_debt,
            "mkt_cap_mm":      mkt_cap,
            "ev_mm":           ev,

            # Cash Flow / DCF
            "capex_mm":        capex,
            "fcf_mm":          fcf,
            "op_cashflow_mm":  op_cf,
            "da_mm":           da,
            "capex_pct_rev":   round(abs(capex) / revenue * 100, 1) if capex and revenue else None,
            "fcf_margin_pct":  round(fcf / revenue * 100, 1) if fcf and revenue else None,

            # WACC inputs
            "beta":            beta,
            "shares_mm":       shares,
            "price":           price,
            "cost_of_debt_pct":  cost_of_debt,
            "equity_weight_pct": equity_weight,
            "debt_weight_pct":   debt_weight,

            # Leverage
            "net_debt_ebitda": net_debt_ebitda,

            # Múltiplos
            "ev_revenue":      round(ev / revenue, 1) if ev and revenue else None,
            "ev_ebitda":       round(ev / ebitda, 1)  if ev and ebitda else None,
            "pe":              round(info.get("trailingPE"), 1) if info.get("trailingPE") else None,
            "pb":              round(info.get("priceToBook"), 2) if info.get("priceToBook") else None,
            "rev_growth_pct":  round(info.get("revenueGrowth") * 100, 1) if info.get("revenueGrowth") else None,
        }

    except Exception as e:
        return {"ticker": ticker, "error": str(e)}


def calcular_wacc(dcf_data: dict, risk_free_rate: float = 4.5, equity_risk_premium: float = 5.5) -> dict:
    """
    Calcula WACC usando CAPM para cost of equity.
    
    Defaults 2025:
    - Risk free rate: 4.5% (US 10Y Treasury)
    - ERP: 5.5% (Damodaran estimate)
    """
    beta            = dcf_data.get("beta")
    cost_of_debt    = dcf_data.get("cost_of_debt_pct")
    tax_rate        = dcf_data.get("tax_rate_pct")
    equity_weight   = dcf_data.get("equity_weight_pct")
    debt_weight     = dcf_data.get("debt_weight_pct")

    result = {
        "risk_free_rate_pct":      risk_free_rate,
        "equity_risk_premium_pct": equity_risk_premium,
        "beta":                    beta,
        "cost_of_equity_pct":      None,
        "cost_of_debt_pct":        cost_of_debt,
        "tax_rate_pct":            tax_rate,
        "after_tax_cost_debt_pct": None,
        "equity_weight_pct":       equity_weight,
        "debt_weight_pct":         debt_weight,
        "wacc_pct":                None,
        "nota":                    "",
    }

    # Cost of Equity = Rf + Beta * ERP
    if beta:
        cost_of_equity = risk_free_rate + beta * equity_risk_premium
        result["cost_of_equity_pct"] = round(cost_of_equity, 2)

    # After-tax cost of debt
    if cost_of_debt and tax_rate:
        after_tax_cod = cost_of_debt * (1 - tax_rate / 100)
        result["after_tax_cost_debt_pct"] = round(after_tax_cod, 2)

    # WACC
    if (result["cost_of_equity_pct"] and result["after_tax_cost_debt_pct"]
            and equity_weight and debt_weight):
        wacc = (result["cost_of_equity_pct"] * equity_weight / 100 +
                result["after_tax_cost_debt_pct"] * debt_weight / 100)
        result["wacc_pct"] = round(wacc, 2)
    elif result["cost_of_equity_pct"] and not cost_of_debt:
        # Sin deuda o no disponible → WACC ≈ Cost of Equity
        result["wacc_pct"] = result["cost_of_equity_pct"]
        result["nota"] = "Cost of debt no disponible; WACC estimado como cost of equity"

    return result


# ─────────────────────────────────────────────
# 5. ESTILOS
# ─────────────────────────────────────────────

NAVY   = "00205B"
DELOITTE = "86BC25"
LBLUE  = "DCE6F1"
GRAY   = "F5F5F5"
WHITE  = "FFFFFF"
YELLOW = "FFFF00"

def hdr_font():   return Font(name="Arial", bold=True, color="FFFFFF", size=9)
def dat_font():   return Font(name="Arial", size=9)
def bold_font():  return Font(name="Arial", bold=True, size=9)
def blue_font():  return Font(name="Arial", size=9, color="0000FF")
def title_font(): return Font(name="Arial", bold=True, size=13, color=DELOITTE)
def sub_font():   return Font(name="Arial", size=9, italic=True, color="666666")
def hdr_fill():   return PatternFill("solid", start_color=DELOITTE)
def sum_fill():   return PatternFill("solid", start_color=LBLUE)
def alt_fill(i):  return PatternFill("solid", start_color=GRAY if i % 2 == 0 else WHITE)

thin = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
AL = Alignment(horizontal="left",   vertical="center")
AR = Alignment(horizontal="right",  vertical="center")
AC = Alignment(horizontal="center", vertical="center")
AW = Alignment(horizontal="left",   vertical="top", wrap_text=True)

# ─────────────────────────────────────────────
# 6. COLUMNAS Y HELPERS
# ─────────────────────────────────────────────

COLS = [
    ("Ticker",        9,  "txt"),
    ("Empresa",       30, "txt"),
    ("País",          14, "txt"),
    ("Sector",        18, "txt"),
    ("Industria",     24, "txt"),
    ("Descripción",   60, "wrap"),
    ("Revenue ($mm)", 15, "num"),
    ("EBITDA ($mm)",  14, "num"),
    ("Net Inc ($mm)", 14, "num"),
    ("Gross ($mm)"),
    ("Deuda ($mm)",   13, "num"),
    ("Cash ($mm)",    13, "num"),
    ("Mkt Cap ($mm)", 15, "num"),
    ("EV ($mm)",      14, "num"),
    ("P/E",           9,  "mult"),
    ("Rev Growth %",  13, "pct"),
    ("Empleados",     13, "int"),
    ("EBITDA Mg%",    13, "fpct"),
    ("Net Mg%",       12, "fpct"),
    ("Gross Mg%",     13, "fpct"),
    ("EV/Revenue",    13, "fmult"),
    ("EV/EBITDA",     13, "fmult"),
]

COLS_VISIBLE = [c for c in COLS if len(c) == 3]


def col_idx(key):
    for i, c in enumerate(COLS_VISIBLE, 1):
        if c[0] == key:
            return i
    return None


def write_header_row(ws, row, cols):
    for ci, (name, width, _) in enumerate(cols, 1):
        cell = ws.cell(row=row, column=ci, value=name)
        cell.font      = hdr_font()
        cell.fill      = hdr_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[row].height = 30


def write_data_row(ws, er, row_data, cols, col_idx_map):
    fill = alt_fill(er)

    for ci, (key, width, fmt) in enumerate(cols, 1):
        val  = row_data.get(key)
        cell = ws.cell(row=er, column=ci, value=val)
        cell.fill   = fill
        cell.border = thin

        if fmt == "wrap":
            cell.font = dat_font(); cell.alignment = AW
        elif fmt == "num":
            cell.font = dat_font(); cell.alignment = AR; cell.number_format = '#,##0.0'
        elif fmt == "mult":
            cell.font = dat_font(); cell.alignment = AR; cell.number_format = '0.0"x"'
        elif fmt == "pct":
            cell.font = dat_font(); cell.alignment = AR; cell.number_format = '0.0"%"'
        elif fmt == "int":
            cell.font = dat_font(); cell.alignment = AR; cell.number_format = '#,##0'
        elif fmt in ("fpct", "fmult"):
            cell.value = None
        else:
            cell.font = dat_font(); cell.alignment = AL

    def cl(key): return get_column_letter(col_idx_map[key])

    rev  = f"{cl('Revenue ($mm)')}{er}"
    ebit = f"{cl('EBITDA ($mm)')}{er}"
    neti = f"{cl('Net Inc ($mm)')}{er}"
    ev   = f"{cl('EV ($mm)')}{er}"

    gross_val = row_data.get("Gross ($mm)")

    formula_map = {
        "EBITDA Mg%": (f"=IF({rev}>0,{ebit}/{rev}*100,\"\")", '0.0"%"'),
        "Net Mg%":    (f"=IF({rev}>0,{neti}/{rev}*100,\"\")", '0.0"%"'),
        "EV/Revenue": (f"=IF({rev}>0,{ev}/{rev},\"\")",       '0.0"x"'),
        "EV/EBITDA":  (f"=IF({ebit}>0,{ev}/{ebit},\"\")",     '0.0"x"'),
    }

    if "Gross ($mm)" in col_idx_map:
        gross_ref = f"{cl('Gross ($mm)')}{er}"
        formula_map["Gross Mg%"] = (f"=IF({rev}>0,{gross_ref}/{rev}*100,\"\")", '0.0"%"')
    else:
        if gross_val and row_data.get("Revenue ($mm)") and row_data["Revenue ($mm)"] > 0:
            formula_map["Gross Mg%"] = (round(gross_val / row_data["Revenue ($mm)"] * 100, 1), '0.0"%"')
        else:
            formula_map["Gross Mg%"] = ("", '0.0"%"')

    for key, (formula, fmt_str) in formula_map.items():
        if key not in col_idx_map:
            continue
        cell = ws.cell(row=er, column=col_idx_map[key], value=formula)
        cell.fill = fill; cell.border = thin
        cell.font = dat_font(); cell.alignment = AR
        cell.number_format = fmt_str


# ─────────────────────────────────────────────
# 7. GENERAR EXCEL — COMPS
# ─────────────────────────────────────────────

def generar_excel(df: pd.DataFrame):
    wb = Workbook()
    df_sorted = df.sort_values("Revenue ($mm)", ascending=False).reset_index(drop=True)
    col_idx_map = {c[0]: i for i, c in enumerate(COLS_VISIBLE, 1)}

    NUM_STAT_COLS = [
        ("Revenue ($mm)", '#,##0.0'), ("EBITDA ($mm)", '#,##0.0'),
        ("Net Inc ($mm)", '#,##0.0'), ("EV ($mm)", '#,##0.0'),
        ("Mkt Cap ($mm)", '#,##0.0'), ("P/E", '0.0"x"'),
        ("Rev Growth %", '0.0"%"'),
        ("EBITDA Mg%", '0.0"%"'), ("Net Mg%", '0.0"%"'), ("Gross Mg%", '0.0"%"'),
        ("EV/Revenue", '0.0"x"'), ("EV/EBITDA", '0.0"x"'),
    ]

    # ════════════════════════
    # HOJA 1: UNIVERSE COMPLETO
    # ════════════════════════
    ws1 = wb.active
    ws1.title = "Universe Completo"
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = "A4"

    ncols = len(COLS_VISIBLE)
    ws1.merge_cells(f"A1:{get_column_letter(ncols)}1")
    ws1["A1"] = f"Comparable Companies — Universe | {DEAL_CONFIG['empresa_target']} | {DEAL_CONFIG['sector']}"
    ws1["A1"].font = title_font()
    ws1.row_dimensions[1].height = 26

    ws1.merge_cells(f"A2:{get_column_letter(ncols)}2")
    ws1["A2"] = f"Analista: {DEAL_CONFIG['analista']}  |  Fecha: {DEAL_CONFIG['fecha']}  |  Fuente: Yahoo Finance  |  {len(df_sorted)} empresas"
    ws1["A2"].font = sub_font()
    ws1.row_dimensions[2].height = 16

    write_header_row(ws1, 4, COLS_VISIBLE)

    DATA_START1 = 5
    for i, row_data in df_sorted.iterrows():
        er = DATA_START1 + i
        ws1.row_dimensions[er].height = 32 if row_data.get("Descripción") else 18
        write_data_row(ws1, er, row_data, COLS_VISIBLE, col_idx_map)

    last1   = DATA_START1 + len(df_sorted) - 1
    stats1  = last1 + 2

    for offset, label in enumerate(["Mediana", "Promedio"], 1):
        sr = stats1 + offset
        ws1.cell(row=sr, column=1, value=label).font = bold_font()
        ws1.cell(row=sr, column=1).fill   = sum_fill()
        ws1.cell(row=sr, column=1).border = thin
        fn = "MEDIAN" if label == "Mediana" else "AVERAGE"
        for key, fmt_str in NUM_STAT_COLS:
            if key not in col_idx_map:
                continue
            ci = col_idx_map[key]
            cl = get_column_letter(ci)
            cell = ws1.cell(row=sr, column=ci, value=f"={fn}({cl}{DATA_START1}:{cl}{last1})")
            cell.fill = sum_fill(); cell.border = thin
            cell.font = bold_font(); cell.alignment = AR
            cell.number_format = fmt_str

    # ════════════════════════
    # HOJA 2: FILTRADAS
    # ════════════════════════
    ws2 = wb.create_sheet("Comps Filtradas")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A8"

    ws2["A1"] = "PARÁMETROS DEL DEAL"
    ws2["A1"].font = Font(name="Arial", bold=True, size=11, color=DELOITTE)
    ws2.row_dimensions[1].height = 20

    param_rows = [
        (2, "Revenue Target ($mm)", DEAL_CONFIG["revenue_target"],    '#,##0'),
        (3, "Rango mínimo (%)",     DEAL_CONFIG["rango_min_pct"]*100, '0"%"'),
        (4, "Rango máximo (%)",     DEAL_CONFIG["rango_max_pct"]*100, '0"%"'),
    ]
    for r_num, label, val, fmt_str in param_rows:
        ws2.cell(row=r_num, column=1, value=label).font = dat_font()
        cell = ws2.cell(row=r_num, column=2, value=val)
        cell.font = blue_font()
        cell.fill = PatternFill("solid", start_color=YELLOW)
        cell.number_format = fmt_str
        cell.border = thin

    ws2.merge_cells(f"A6:{get_column_letter(ncols)}6")
    ws2["A6"] = "Empresas con revenue entre $B$2*(B3/100) y $B$2*(B4/100) — modificá B2:B4 para cambiar el filtro"
    ws2["A6"].font = sub_font()
    ws2.row_dimensions[6].height = 14
    ws2.row_dimensions[5].height = 8

    write_header_row(ws2, 7, COLS_VISIBLE)

    rev_min = DEAL_CONFIG["revenue_target"] * DEAL_CONFIG["rango_min_pct"]
    rev_max = DEAL_CONFIG["revenue_target"] * DEAL_CONFIG["rango_max_pct"]
    df_filt = df_sorted[
        (df_sorted["Revenue ($mm)"].notna()) &
        (df_sorted["Revenue ($mm)"] >= rev_min) &
        (df_sorted["Revenue ($mm)"] <= rev_max)
    ].reset_index(drop=True)

    DATA_START2 = 8
    for i, row_data in df_filt.iterrows():
        er = DATA_START2 + i
        ws2.row_dimensions[er].height = 42
        write_data_row(ws2, er, row_data, COLS_VISIBLE, col_idx_map)

    last2   = DATA_START2 + len(df_filt) - 1
    med_row = last2 + 2
    avg_row = last2 + 3

    for offset, label in enumerate(["Mediana", "Promedio"], 1):
        sr = last2 + 1 + offset
        ws2.cell(row=sr, column=1, value=label).font = bold_font()
        ws2.cell(row=sr, column=1).fill   = sum_fill()
        ws2.cell(row=sr, column=1).border = thin
        fn = "MEDIAN" if label == "Mediana" else "AVERAGE"
        for key, fmt_str in NUM_STAT_COLS:
            if key not in col_idx_map:
                continue
            ci = col_idx_map[key]
            cl = get_column_letter(ci)
            cell = ws2.cell(row=sr, column=ci, value=f"={fn}({cl}{DATA_START2}:{cl}{last2})")
            cell.fill = sum_fill(); cell.border = thin
            cell.font = bold_font(); cell.alignment = AR
            cell.number_format = fmt_str

    # IMPLIED VALUATION
    impl = last2 + 6
    ws2.cell(row=impl, column=1, value="── IMPLIED VALUATION ──").font = Font(name="Arial", bold=True, size=10, color=DELOITTE)
    ws2.row_dimensions[impl].height = 20

    ev_rev_ci = col_idx_map.get("EV/Revenue")
    ev_ebt_ci = col_idx_map.get("EV/EBITDA")
    ebt_mg_ci = col_idx_map.get("EBITDA Mg%")

    ev_rev_ref = f"{get_column_letter(ev_rev_ci)}{med_row}"
    ev_ebt_ref = f"{get_column_letter(ev_ebt_ci)}{med_row}"
    ebt_mg_ref = f"{get_column_letter(ebt_mg_ci)}{med_row}"

    impl_rows = [
        ("Revenue Target ($mm)",             "=$B$2",                                        '#,##0'),
        ("EV Implied — EV/Revenue (mediana)", f"=$B$2*{ev_rev_ref}",                          '#,##0'),
        ("EV Implied — EV/EBITDA (mediana)",  f"=$B$2*({ebt_mg_ref}/100)*{ev_ebt_ref}",       '#,##0'),
    ]
    for j, (lbl, formula, fmt_str) in enumerate(impl_rows):
        ws2.cell(row=impl+1+j, column=1, value=lbl).font = dat_font()
        cell = ws2.cell(row=impl+1+j, column=2, value=formula)
        cell.font = bold_font(); cell.fill = sum_fill()
        cell.border = thin; cell.alignment = AR
        cell.number_format = fmt_str

    fecha_str = datetime.now().strftime("%Y%m%d")
    fname = f"Comps_{DEAL_CONFIG['empresa_target'].replace(' ','_')}_{fecha_str}.xlsx"
    wb.save(fname)
    print(f"\n  ✅ Excel guardado: {fname}")
    return fname


def _generar_excel_buffer(df: pd.DataFrame, buffer) -> None:
    """Igual que generar_excel pero guarda en un BytesIO en vez de disco.
    Usa el DEAL_CONFIG que ya fue seteado antes de llamarla."""
    wb = Workbook()
    df_sorted = df.sort_values("Revenue ($mm)", ascending=False).reset_index(drop=True)
    col_idx_map = {c[0]: i for i, c in enumerate(COLS_VISIBLE, 1)}

    NUM_STAT_COLS = [
        ("Revenue ($mm)", '#,##0.0'), ("EBITDA ($mm)", '#,##0.0'),
        ("Net Inc ($mm)", '#,##0.0'), ("EV ($mm)", '#,##0.0'),
        ("Mkt Cap ($mm)", '#,##0.0'), ("P/E", '0.0"x"'),
        ("Rev Growth %", '0.0"%"'),
        ("EBITDA Mg%", '0.0"%"'), ("Net Mg%", '0.0"%"'), ("Gross Mg%", '0.0"%"'),
        ("EV/Revenue", '0.0"x"'), ("EV/EBITDA", '0.0"x"'),
    ]

    # ── HOJA 1: UNIVERSE COMPLETO ──
    ws1 = wb.active
    ws1.title = "Universe Completo"
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = "A4"
    ncols = len(COLS_VISIBLE)

    ws1.merge_cells(f"A1:{get_column_letter(ncols)}1")
    ws1["A1"] = f"Comparable Companies — Universe | {DEAL_CONFIG['empresa_target']} | {DEAL_CONFIG['sector']}"
    ws1["A1"].font = title_font()
    ws1.row_dimensions[1].height = 26

    ws1.merge_cells(f"A2:{get_column_letter(ncols)}2")
    ws1["A2"] = f"Analista: {DEAL_CONFIG['analista']}  |  Fecha: {DEAL_CONFIG['fecha']}  |  Fuente: Yahoo Finance  |  {len(df_sorted)} empresas"
    ws1["A2"].font = sub_font()
    ws1.row_dimensions[2].height = 16
    write_header_row(ws1, 3, COLS_VISIBLE)
    DATA_START1 = 4
    for i, row_data in df_sorted.iterrows():
        er = DATA_START1 + i
        ws1.row_dimensions[er].height = 32 if row_data.get("Descripción") else 18
        write_data_row(ws1, er, row_data, COLS_VISIBLE, col_idx_map)

    last1 = DATA_START1 + len(df_sorted) - 1
    stats1 = last1 + 2
    for offset, label in enumerate(["Mediana", "Promedio"], 1):
        sr = stats1 + offset
        ws1.cell(row=sr, column=1, value=label).font = bold_font()
        ws1.cell(row=sr, column=1).fill   = sum_fill()
        ws1.cell(row=sr, column=1).border = thin
        fn = "MEDIAN" if label == "Mediana" else "AVERAGE"
        for key, fmt_str in NUM_STAT_COLS:
            if key not in col_idx_map: continue
            ci = col_idx_map[key]
            cl = get_column_letter(ci)
            cell = ws1.cell(row=sr, column=ci, value=f"={fn}({cl}{DATA_START1}:{cl}{last1})")
            cell.fill = sum_fill(); cell.border = thin
            cell.font = bold_font(); cell.alignment = AR
            cell.number_format = fmt_str

    # ── HOJA 2: FILTRADAS ──
    ws2 = wb.create_sheet("Comps Filtradas")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A8"

    ws2["A1"] = "PARÁMETROS DEL DEAL"
    ws2["A1"].font = Font(name="Arial", bold=True, size=11, color=DELOITTE)
    ws2.row_dimensions[1].height = 20

    param_rows = [
        (2, "Revenue Target ($mm)", DEAL_CONFIG["revenue_target"],    '#,##0'),
        (3, "Rango mínimo (%)",     DEAL_CONFIG["rango_min_pct"]*100, '0"%"'),
        (4, "Rango máximo (%)",     DEAL_CONFIG["rango_max_pct"]*100, '0"%"'),
    ]
    for r_num, label, val, fmt_str in param_rows:
        ws2.cell(row=r_num, column=1, value=label).font = dat_font()
        cell = ws2.cell(row=r_num, column=2, value=val)
        cell.font = blue_font(); cell.fill = PatternFill("solid", start_color=YELLOW)
        cell.number_format = fmt_str; cell.border = thin

    ws2.merge_cells(f"A6:{get_column_letter(ncols)}6")
    ws2["A6"] = "Empresas con revenue entre $B$2*(B3/100) y $B$2*(B4/100) — modificá B2:B4 para cambiar el filtro"
    ws2["A6"].font = sub_font()
    ws2.row_dimensions[6].height = 14
    ws2.row_dimensions[5].height = 8

    write_header_row(ws2, 7, COLS_VISIBLE)

    rev_min = DEAL_CONFIG["revenue_target"] * DEAL_CONFIG["rango_min_pct"]
    rev_max = DEAL_CONFIG["revenue_target"] * DEAL_CONFIG["rango_max_pct"]
    df_filt = df_sorted[
        (df_sorted["Revenue ($mm)"].notna()) &
        (df_sorted["Revenue ($mm)"] >= rev_min) &
        (df_sorted["Revenue ($mm)"] <= rev_max)
    ].reset_index(drop=True)

    DATA_START2 = 8
    for i, row_data in df_filt.iterrows():
        er = DATA_START2 + i
        ws2.row_dimensions[er].height = 42
        write_data_row(ws2, er, row_data, COLS_VISIBLE, col_idx_map)

    last2 = DATA_START2 + len(df_filt) - 1
    med_row = last2 + 2
    for offset, label in enumerate(["Mediana", "Promedio"], 1):
        sr = last2 + 1 + offset
        ws2.cell(row=sr, column=1, value=label).font = bold_font()
        ws2.cell(row=sr, column=1).fill   = sum_fill()
        ws2.cell(row=sr, column=1).border = thin
        fn = "MEDIAN" if label == "Mediana" else "AVERAGE"
        for key, fmt_str in NUM_STAT_COLS:
            if key not in col_idx_map: continue
            ci = col_idx_map[key]
            cl = get_column_letter(ci)
            cell = ws2.cell(row=sr, column=ci, value=f"={fn}({cl}{DATA_START2}:{cl}{last2})")
            cell.fill = sum_fill(); cell.border = thin
            cell.font = bold_font(); cell.alignment = AR
            cell.number_format = fmt_str

    # IMPLIED VALUATION
    ev_rev_ci = col_idx_map.get("EV/Revenue")
    ev_ebt_ci = col_idx_map.get("EV/EBITDA")
    ebt_mg_ci = col_idx_map.get("EBITDA Mg%")
    ev_rev_ref = f"{get_column_letter(ev_rev_ci)}{med_row}"
    ev_ebt_ref = f"{get_column_letter(ev_ebt_ci)}{med_row}"
    ebt_mg_ref = f"{get_column_letter(ebt_mg_ci)}{med_row}"

    impl = last2 + 6
    ws2.cell(row=impl, column=1, value="── IMPLIED VALUATION ──").font = Font(name="Arial", bold=True, size=10, color=DELOITTE)
    ws2.row_dimensions[impl].height = 20

    impl_rows = [
        ("Revenue Target ($mm)",             "=$B$2",                                       '#,##0'),
        ("EV Implied — EV/Revenue (mediana)", f"=$B$2*{ev_rev_ref}",                         '#,##0'),
        ("EV Implied — EV/EBITDA (mediana)",  f"=$B$2*({ebt_mg_ref}/100)*{ev_ebt_ref}",      '#,##0'),
    ]
    for j, (lbl, formula, fmt_str) in enumerate(impl_rows):
        ws2.cell(row=impl+1+j, column=1, value=lbl).font = dat_font()
        cell = ws2.cell(row=impl+1+j, column=2, value=formula)
        cell.font = bold_font(); cell.fill = sum_fill()
        cell.border = thin; cell.alignment = AR
        cell.number_format = fmt_str

    wb.save(buffer)


# ─────────────────────────────────────────────
# 8. MAIN
# ─────────────────────────────────────────────

def run():
    print("\n" + "="*60)
    print(f"  📊 M&A COMPS v3 — {DEAL_CONFIG['empresa_target']}")
    print(f"  Sector: {DEAL_CONFIG['sector']}  |  Revenue target: ${DEAL_CONFIG['revenue_target']:,}mm")
    print("="*60)

    tickers = UNIVERSE.get(DEAL_CONFIG["sector"], UNIVERSE["Health Insurance"])
    print(f"\n  Descargando datos de {len(tickers)} empresas...\n")

    resultados = []
    for i, ticker in enumerate(tickers):
        print(f"  [{i+1:02d}/{len(tickers)}] {ticker:<12}", end=" ")
        data = get_financials(ticker)
        if data and data.get("Revenue ($mm)"):
            resultados.append(data)
            print(f"✅  ${data['Revenue ($mm)']:>10,.0f}mm  {data.get('Empresa','')}")
        else:
            print("⚠️  sin datos")
        time.sleep(0.25)

    if not resultados:
        print("\n  ❌ Sin datos.")
        return

    df = pd.DataFrame(resultados)
    generar_excel(df)


if __name__ == "__main__":
    run()
